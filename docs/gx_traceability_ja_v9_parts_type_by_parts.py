#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GeneXus Traceability Builder v9 (hierarchy + parts Source)
- Stage 1: Parse GeneXus design exports (XML/JSON/text/XPZ/ZIP) -> build design object inventory (with GUID hierarchy) ->
           infer feature names -> Feature_Design / Feature_Group
- Stage 2: Scan generated Java sources -> map design object -> Java files -> Design_Java
Output: Excel (.xlsx) with 3 sheets

v5重点改动（用于你的场景）：
1) 解析GeneXus导出中 Object 的层次信息：guid / parentGuid / moduleGuid / type(通常为GUID) / fullyQualifiedName
2) 支持“通过对象查找定义”：
   - 同一 guid 可能在不同文件出现（有的只有type GUID，有的有typeName/ObjectType），v5会按 guid 合并信息
   - 自动构建 type_guid -> type_name 映射，用于补全只有type GUID的对象类型
3) Source提取增强：
   - 不只取第一个 <Source>，会收集 Object 下所有 Part 子项中的所有 <Source>，以及 Property(Name=Source*) 的 Value
   - 去重后按出现顺序拼接，控制长度避免Excel过大
4) 修复“全部被判成WP”的类型推断：
   - WP不再使用容易误判的短关键字 "wp"
   - 同分时更偏向更具体类型（DP/Trn/SDT/Proc/WWP/WP）
"""

import argparse
import json
import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
except Exception as e:  # pragma: no cover
    raise SystemExit("openpyxl is required. Please install: pip install openpyxl") from e


# -----------------------------
# Utilities
# -----------------------------

UUID_RE = re.compile(r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$")


# -----------------------------
# GeneXus internal object-type GUID mapping
# NOTE: This mapping can vary slightly by GeneXus generation, but these suffixes are commonly observed.
# We match by the last 4 digits of the final GUID block (12 digits).
GX_TYPE_GUID_SUFFIX_MAP = {
    # NOTE: 末尾4桁による種別推定は世代差でズレる可能性があるため、
    # ここでは画像の指針で明示された Folder/Category（…0006）のみ特判定する。
    "0006": ("Folder", "Folder / Category"),
}

def map_gx_type_guid(type_guid: str):
    """Map a GeneXus object 'type' GUID to (normalized_type, display_name)."""
    g = clean(type_guid)
    if not g or not UUID_RE.match(g):
        return None
    last = g.split("-")[-1]  # 12 chars
    suffix = last[-4:] if len(last) >= 4 else last
    return GX_TYPE_GUID_SUFFIX_MAP.get(suffix)



def is_uuid(s: str) -> bool:
    return bool(s and UUID_RE.match(s.strip()))


def clean(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def read_text_safely(path: Path, max_bytes: int = 2_000_000) -> str:
    """Read file with encoding fallback. Limit bytes to avoid huge files."""
    data = path.read_bytes()
    if len(data) > max_bytes:
        data = data[:max_bytes]
    for enc in ("utf-8", "utf-8-sig", "cp932", "shift_jis", "gbk", "latin-1"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="replace")


def safe_extract_zip(zf: zipfile.ZipFile, out_dir: Path) -> None:
    """Prevent zip slip by validating output paths."""
    out_dir = out_dir.resolve()
    for member in zf.infolist():
        member_path = (out_dir / member.filename).resolve()
        if not str(member_path).startswith(str(out_dir)):
            continue
        if member.is_dir():
            member_path.mkdir(parents=True, exist_ok=True)
        else:
            member_path.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(member, "r") as src, open(member_path, "wb") as dst:
                dst.write(src.read())


# -----------------------------
# Data model
# -----------------------------

@dataclass
class DesignObject:
    object_name: str
    object_type: str
    source_file: str

    guid: str = ""
    parent_guid: str = ""
    module_guid: str = ""
    fully_qualified_name: str = ""
    raw_type_id: str = ""     # often a GUID
    raw_type_name: str = ""   # e.g., WebPanel / Transaction / DataProvider, if available

    source: str = ""          # extracted Source snippets (from Parts too)
    parts: Dict[str, str] = field(default_factory=dict)  # extracted Parts content by tag name
    parts_all: str = ""  # concatenated raw <Part> payload (truncated)
    title: str = ""
    description: str = ""
    raw_hints: str = ""
    inferred_entity: str = ""
    inferred_actions: str = ""
    feature_names: str = ""


# -----------------------------
# Object type normalize & guess
# -----------------------------

TYPE_ALIASES = {
    # Common
    "webpanel": "WP",
    "web panel": "WP",
    "web_panel": "WP",
    "wp": "WP",
    "workpanel": "WP",

    "workwithplus": "WWP",
    "work with plus": "WWP",
    "workwith": "WWP",
    "wwp": "WWP",

    "transaction": "Transaction",
    "trn": "Transaction",

    "procedure": "Procedure",
    "proc": "Procedure",

    "dataprovider": "DP",
    "data provider": "DP",
    "dp": "DP",

    "structureddata": "SDT",
    "structureddatatype": "SDT",
    "sdt": "SDT",

    "sd": "SD",
    "smartdevicepanel": "SD",
    "smart device panel": "SD",

    "rest": "API",
    "api": "API",
    "theme": "Theme",
    "thm": "Theme",
    "service": "API",

    "gam": "GAM",
    "gam object": "GAM",
    "security": "GAM",
}


def normalize_object_type(raw: str, filename: str = "", object_name: str = "") -> str:
    """Normalize raw type name (or id string) to a stable category."""
    r = clean(raw).lower()
    if r in TYPE_ALIASES:
        return TYPE_ALIASES[r]

    mapped = map_gx_type_guid(raw)
    if mapped:
        return mapped[0]

    # If raw is a GUID, we cannot normalize directly. Try filename/object_name heuristics.
    name = (object_name or "")
    fn = (filename or "").lower()

    # Strong filename/name patterns (high precision)
    if re.search(r"(_dp\d*$|_dp_|dataprovider)", name.lower()) or "_dp" in fn:
        return "DP"
    if re.search(r"(_trn\d*$|trn$|transaction)", name.lower()) or "_trn" in fn:
        return "Transaction"
    if re.search(r"(sdt$|_sdt_|structureddatatype)", name.lower()) or "_sdt" in fn:
        return "SDT"
    if re.search(r"(proc$|procedure)", name.lower()) or "_proc" in fn:
        return "Procedure"
    if re.search(r"(ww$|wwp|workwith)", name.lower()) or "workwith" in fn:
        return "WWP"

    # Default
    return "WP" if raw else "Unknown"


# For guessing types from text/name when no reliable raw type name exists
# IMPORTANT: do NOT include a bare "wp" substring keyword (too many false positives).
TYPE_KEYWORDS_ORDERED = [
    ("DP", [r"\bdataprovider\b", r"\bdp\b", r"_dp\d*\b"]),
    ("Transaction", [r"\btransaction\b", r"\btrn\b", r"_trn\d*\b"]),
    ("SDT", [r"\bsdt\b", r"\bstructureddatatype\b", r"_sdt\b"]),
    ("Procedure", [r"\bprocedure\b", r"\bproc\b", r"_proc\b"]),
    ("WWP", [r"\bworkwith\b", r"\bwwp\b", r"\bww\b"]),
    ("SD", [r"\bsmart\s*device\b", r"\bsd\b", r"\bpanel\b"]),
    ("API", [r"\brest\b", r"\bapi\b", r"\bservice\b"]),
    ("WP", [r"\bweb\s*panel\b", r"\bwebpanel\b"]),
]



def _canonicalize_parts(parts: Optional[Dict[str, str]]) -> Dict[str, str]:
    """Part辞書のキーを正規化（Source/Rules/...）。
    正規XML解析と正規表現フォールバックでキーの大小がブレるため、出力と判定を安定させる。
    """
    out: Dict[str, str] = {}
    if not parts:
        return out
    canon = {t.lower(): t for t in PART_TAGS_ORDER}
    for k, v in parts.items():
        kk = clean(str(k))
        if not kk:
            continue
        key = canon.get(kk.lower(), kk)  # 未知タグはそのまま保持
        vv = (v or "").strip()
        if not vv:
            continue
        if key in out and out[key]:
            norm_existing = re.sub(r"\s+", " ", out[key]).strip()
            norm_new = re.sub(r"\s+", " ", vv).strip()
            if norm_new and norm_new != norm_existing:
                out[key] = out[key] + "\n\n---\n\n" + vv
        else:
            out[key] = vv
    return out


def _type_guid_is_folder(type_guid: str) -> bool:
    g = clean(type_guid)
    if not g or not is_uuid(g):
        return False
    last = g.split("-")[-1]
    suf = last[-4:] if len(last) >= 4 else last
    return suf == "0006"


def infer_object_type_from_parts(parts: Dict[str, str], parts_all: str, text_sample: str, raw_type_id: str) -> str:
    """画像の指針（Part構成）から ObjectType を推定する（核心）。
    ルール（優先順）:
      - type=…0006 → Folder/Category
      - Security / GAM Property → GAM Object
      - Structure + Attributes + Rules → Transaction
      - Source + Parm + Variables → Procedure
      - Source + Output Structure → Data Provider
      - Events + Layout + WWP系Property → WWP
      - Events + Layout → Web Panel(WP)
      - Structureのみ → SDT
    """
    parts = _canonicalize_parts(parts)
    present = {k for k in PART_TAGS_ORDER if (parts.get(k, "") or "").strip()}

    blob = ((parts_all or "") + "\n" + (text_sample or "")).lower()

    # Folder / Category
    if _type_guid_is_folder(raw_type_id):
        return "Folder"

    # GAM / Security
    if re.search(r"<name>[^<]*(gam|security)[^<]*</name>", blob) or re.search(r"\bgam\b", blob):
        return "GAM"

    # Transaction
    if {"Structure", "Attributes", "Rules"}.issubset(present):
        return "Transaction"

    # Procedure
    if {"Source", "Parm", "Variables"}.issubset(present):
        return "Procedure"

    # Data Provider: Source + Output Structure
    if "Source" in present:
        if re.search(r"<output\b", blob) or "outputstructure" in blob or re.search(r"output\s+structure", blob):
            return "DP"

    # SDT: Structure only
    if "Structure" in present:
        other = present - {"Structure"}
        if not other:
            return "SDT"

    # Web Panel / WWP (both have Events+Layout; WWP has WWP系Property)
    if {"Events", "Layout"}.issubset(present):
        if re.search(r"workwithplus", blob) or re.search(r"<name>[^<]*wwp[^<]*</name>", blob):
            return "WWP"
        return "WP"

    return ""


def guess_type_from_text(text: str, object_name: str, filename: str) -> str:
    t = (text or "").lower()
    name = (object_name or "").lower()
    fn = (filename or "").lower()
    best_type = "Unknown"
    best_score = -1

    for typ, patterns in TYPE_KEYWORDS_ORDERED:
        score = 0
        for pat in patterns:
            if re.search(pat, t) or re.search(pat, name) or re.search(pat, fn):
                score += 1
        if score > best_score:
            best_score = score
            best_type = typ

    return best_type if best_score > 0 else "Unknown"


# -----------------------------
# XML extraction helpers (hierarchy + parts source)
# -----------------------------

def _iter_prop_nodes(elem: ET.Element) -> List[ET.Element]:
    # Supports: <Properties><Property><Name>..</Name><Value>..</Value></Property></Properties>
    # Also allow <Property name=".." value=".."/>
    props = []
    props.extend(elem.findall(".//Properties/Property"))
    props.extend(elem.findall(".//Property"))
    return props


def extract_properties(elem: ET.Element) -> Dict[str, str]:
    """Extract first occurrence of Name/Value properties into a dict (used for metadata only)."""
    props: Dict[str, str] = {}
    for p in _iter_prop_nodes(elem):
        k = clean(p.attrib.get("name", "") or "")
        v = clean(p.attrib.get("value", "") or "")
        if not k:
            k = clean((p.findtext("Name") or p.findtext("name") or ""))
        if not v:
            v = clean((p.findtext("Value") or p.findtext("value") or ""))
        if k and v and k not in props:
            props[k] = v
    return props


def collect_sources(elem: ET.Element) -> List[str]:
    """Collect all Source snippets under the element:
    1) All <Source>...</Source> text under Object/Part subtree
    2) Any Property with Name == Source or endswith Source (case-insensitive)
    3) Attribute 'source' if present
    Deduplicate, keep order.
    """
    sources: List[str] = []

    # 1) Source tags (possibly many)
    for s in elem.findall(".//Source"):
        txt = ET.tostring(s, encoding="unicode", method="text")
        txt = txt.replace("\r\n", "\n").replace("\r", "\n").strip()
        if txt:
            sources.append(txt)

    # 2) Properties mentioning Source (including parts)
    for p in _iter_prop_nodes(elem):
        k = clean(p.attrib.get("name", "") or "")
        if not k:
            k = clean((p.findtext("Name") or p.findtext("name") or ""))
        if not k:
            continue
        if k.lower() == "source" or k.lower().endswith("source"):
            v = clean(p.attrib.get("value", "") or "")
            if not v:
                v = (p.findtext("Value") or p.findtext("value") or "").strip()
            v = v.replace("\r\n", "\n").replace("\r", "\n").strip()
            if v:
                sources.append(v)

    # 3) Source attributes
    for attr_k, attr_v in elem.attrib.items():
        if attr_k.lower() == "source":
            v = clean(attr_v).replace("\r\n", "\n").replace("\r", "\n").strip()
            if v:
                sources.append(v)

    # Dedup preserve order
    seen = set()
    out: List[str] = []
    for s in sources:
        key = s.strip()
        if not key:
            continue
        if key in seen:
            continue
        seen.add(key)
        out.append(s)
    return out


def join_sources(sources: List[str], limit: int = 10000) -> str:
    if not sources:
        return ""
    blob = "\n\n-----\n\n".join(sources)
    if len(blob) > limit:
        return blob[:limit] + "\n…(truncated)"
    return blob




# -----------------------------
# Part extraction (per Object type)
# -----------------------------
PART_TAGS_ORDER = ["Source", "Rules", "Variables", "Parm", "Events", "Layout", "Attributes", "Structure", "Style"]

# Based on the guideline:
# Procedure: Source / Rules / Variables / Parm
# Web Panel(WP): Events / Rules / Variables / Layout
# WWP: Events / Rules / Variables / Layout
# Transaction: Rules / Events / Attributes / Layout
# Data Provider(DP): Source / Variables / Parm
# REST(API): Source / Parm / Rules
# SDT: Structure
# Theme: Style
TYPE_TO_WRITABLE_PARTS = {
    "Procedure": ["Source", "Rules", "Variables", "Parm"],
    "WP": ["Events", "Rules", "Variables", "Layout"],
    "WWP": ["Events", "Rules", "Variables", "Layout"],
    "SD": ["Events", "Rules", "Variables", "Layout"],   # treat SmartDevicePanel similar to Web Panel
    "Transaction": ["Rules", "Events", "Attributes", "Layout"],
    "DP": ["Source", "Variables", "Parm"],
    "API": ["Source", "Parm", "Rules"],
    "SDT": ["Structure"],
    "Theme": ["Style"],
}

def _local(tag: str) -> str:
    return tag.split("}")[-1] if tag else ""

def _iter_elems_by_localname(elem: ET.Element, localname: str):
    ln = localname.lower()
    for e in elem.iter():
        if _local(e.tag).lower() == ln:
            yield e

def _node_payload(node: ET.Element, limit: int = 20000) -> str:
    """
    Return a payload for a section node (Source/Rules/Events/...):
    - If node has children -> keep XML (to preserve all content)
    - Else -> keep text
    """
    try:
        if list(node):
            s = ET.tostring(node, encoding="unicode")
        else:
            s = ET.tostring(node, encoding="unicode", method="text")
    except Exception:
        s = (node.text or "")
    s = (s or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not s:
        return ""
    if len(s) > limit:
        return s[:limit] + "\n…(truncated)"
    return s

def _join_blobs(blobs: List[str], limit: int = 20000) -> str:
    if not blobs:
        return ""
    blob = "\n\n-----\n\n".join([b for b in blobs if b])
    if len(blob) > limit:
        return blob[:limit] + "\n…(truncated)"
    return blob

def collect_parts_content(obj_elem: ET.Element, object_type_hint: str = "") -> Tuple[Dict[str, str], str]:
    """
    Extract ALL contents under <Part> nodes, and also build per-tag content for:
    Source/Rules/Variables/Parm/Events/Layout/Attributes/Structure/Style

    Returns:
      (parts_by_tag, parts_all_raw)
    """
    # 1) collect raw <Part> XML for "all parts"
    part_xmls: List[str] = []
    part_elems = list(_iter_elems_by_localname(obj_elem, "Part"))
    for pe in part_elems:
        try:
            part_xmls.append(ET.tostring(pe, encoding="unicode"))
        except Exception:
            # fallback: text only
            part_xmls.append(ET.tostring(pe, encoding="unicode", method="text"))
    parts_all = _join_blobs([clean(x) for x in part_xmls if x], limit=40000)

    # 2) collect per-tag contents within each Part subtree
    buckets: Dict[str, List[str]] = {k: [] for k in PART_TAGS_ORDER}
    target_tags = {t.lower(): t for t in PART_TAGS_ORDER}

    for pe in part_elems:
        # tag-based sections
        for node in pe.iter():
            lname = _local(node.tag)
            key = target_tags.get(lname.lower())
            if key:
                payload = _node_payload(node)
                if payload:
                    buckets[key].append(payload)

        # property-based sections (Name == Rules/Events/... etc)
        for pn in _iter_prop_nodes(pe):
            k = clean(pn.attrib.get("name", "") or "")
            if not k:
                k = clean((pn.findtext("Name") or pn.findtext("name") or ""))
            if not k:
                continue
            kl = k.lower()
            for t in PART_TAGS_ORDER:
                tl = t.lower()
                if kl == tl or kl.endswith(tl):
                    v = clean(pn.attrib.get("value", "") or "")
                    if not v:
                        v = (pn.findtext("Value") or pn.findtext("value") or "").strip()
                    v = v.replace("\r\n", "\n").replace("\r", "\n").strip()
                    if v:
                        buckets[t].append(v)

    # 3) dedup & filter by type (only keep "writable parts" defined by the guideline)
    allowed = TYPE_TO_WRITABLE_PARTS.get(object_type_hint or "", PART_TAGS_ORDER)
    parts_by_tag: Dict[str, str] = {}
    for k in PART_TAGS_ORDER:
        if allowed and k not in allowed:
            continue
        items = buckets.get(k, [])
        if not items:
            continue
        seen = set()
        uniq: List[str] = []
        for it in items:
            key = re.sub(r"\s+", " ", it.strip())
            if not key or key in seen:
                continue
            seen.add(key)
            uniq.append(it.strip())
        if uniq:
            parts_by_tag[k] = _join_blobs(uniq, limit=30000)

    return parts_by_tag, parts_all

def merge_parts_dict(a: Optional[Dict[str, str]], b: Optional[Dict[str, str]]) -> Dict[str, str]:
    """Merge two parts dicts by concatenating non-empty values with separators (dedup on normalized content)."""
    out: Dict[str, str] = {}
    for src in (a or {}, b or {}):
        for k, v in (src or {}).items():
            v = (v or "").strip()
            if not v:
                continue
            if k not in out or not out[k]:
                out[k] = v
            else:
                # avoid dup by normalized comparison
                norm_existing = re.sub(r"\s+", " ", out[k]).strip()
                norm_new = re.sub(r"\s+", " ", v).strip()
                if norm_new and norm_new not in norm_existing:
                    out[k] = (out[k].rstrip() + "\n\n-----\n\n" + v).strip()
    return out


def extract_objects_from_xml(xml_text: str, source_label: str) -> List[Dict]:
    """Return list of raw dict records extracted from XML. Supports malformed XML via regex fallback."""
    xml_text = xml_text or ""
    recs: List[Dict] = []

    def from_elem(obj_elem: ET.Element) -> Dict:
        attrs = {k: clean(v) for k, v in obj_elem.attrib.items()}
        props = extract_properties(obj_elem)

        guid = attrs.get("guid", "") or props.get("Guid", "") or props.get("GUID", "")
        parent_guid = attrs.get("parentGuid", "") or attrs.get("parentGUID", "") or props.get("ParentGuid", "")
        module_guid = attrs.get("moduleGuid", "") or props.get("ModuleGuid", "")

        fq = attrs.get("fullyQualifiedName", "") or props.get("FullyQualifiedName", "") or ""
        name = attrs.get("name", "") or attrs.get("Name", "") or props.get("Name", "") or props.get("ObjectName", "") or ""
        # Prefer fully qualified name when it looks real (has dots and longer)
        object_name = fq if fq and (len(fq) >= len(name)) else name
        object_name = clean(object_name)

        raw_type_id = attrs.get("type", "") or attrs.get("objectType", "") or attrs.get("ObjectType", "")
        raw_type_name = attrs.get("typeName", "") or attrs.get("objectTypeName", "") or attrs.get("TypeName", "")
        raw_type_name = raw_type_name or props.get("ObjectType", "") or props.get("Type", "")

        title = clean(props.get("Title", "") or props.get("Caption", "") or attrs.get("title", ""))
        desc = clean(attrs.get("description", "") or props.get("Description", "") or props.get("Desc", ""))

        # hints: keyword scan from subtree text
        elem_text = ET.tostring(obj_elem, encoding="unicode", method="text")
        combined = (elem_text[:20000] + " " + xml_text[:5000]).lower()
        keywords = re.findall(
            r"\b(search|filter|grid|for\s+each|insert|update|delete|new|edit|view|detail|rest|api|dataprovider|sdt|transaction|procedure|dp)\b",
            combined, flags=re.I
        )
        hints = ",".join(sorted(set(k.lower() for k in keywords)))[:300] if keywords else ""
        # mark enum
        if "enumdefinedvalues" in combined:
            hints = (hints + ("," if hints else "") + "enum")[:300]

        # sources (include Parts)
        sources = collect_sources(obj_elem)

        # parts content (ALL contents under <Part> + per-tag content)
        type_hint = ""
        try:
            if raw_type_name:
                type_hint = normalize_object_type(raw_type_name, filename=source_label, object_name=object_name)
            elif raw_type_id and not is_uuid(raw_type_id):
                type_hint = normalize_object_type(raw_type_id, filename=source_label, object_name=object_name)
        except Exception:
            type_hint = ""
        parts_by_tag, parts_all = collect_parts_content(obj_elem, "")  # 先に全Partを抽出（型推定は後段で実施）


        return {
            "guid": guid,
            "parent_guid": parent_guid,
            "module_guid": module_guid,
            "fully_qualified_name": fq,
            "object_name": object_name,
            "raw_type_id": raw_type_id,
            "raw_type_name": raw_type_name,
            "title": title,
            "description": desc,
            "hints": hints,
            "sources": sources,
            "parts": parts_by_tag,
            "parts_all": parts_all,
            "source_file": source_label,
            "text_sample": combined[:5000],
        }

    # Try XML parse first
    try:
        root = ET.fromstring(xml_text)
        # namespace-agnostic search: match localname == 'Object'
        def is_object_tag(tag: str) -> bool:
            return tag.split("}")[-1].lower() == "object"

        obj_elems: List[ET.Element] = []
        if is_object_tag(root.tag):
            obj_elems.append(root)
        obj_elems.extend([e for e in root.iter() if is_object_tag(e.tag)])

        # De-dup by id
        uniq = []
        seen_ids = set()
        for e in obj_elems:
            if id(e) in seen_ids:
                continue
            seen_ids.add(id(e))
            uniq.append(e)

        # The iteration includes root; ensure we don't double-count.
        # Also: some documents contain many nested Object tags; that's fine—we want all.
        for e in uniq:
            rec = from_elem(e)
            if rec.get("object_name") or rec.get("guid") or rec.get("raw_type_id"):
                recs.append(rec)
        if recs:
            return recs
    except Exception:
        pass

    # Regex fallback for malformed "Object header + plain source text"
    # Some GeneXus exports are not well-formed XML (e.g., <Object ...> header followed by plain code),
    # or may contain multiple <Object ...> blocks. In this case, extract ALL Object headers by regex.
    obj_starts = list(re.finditer(r"<Object\b([^>]*)>", xml_text, flags=re.I))
    if obj_starts:
        # Helper: extract blobs for known Part tags from a text segment
        def _extract_part_blobs(seg: str) -> Dict[str, str]:
            parts: Dict[str, str] = {}
            # Common "user editable" parts across object types (per GeneXus)
            for tag in ("Source","Rules","Variables","Parm","Events","Layout","Attributes","Structure","Style"):
                blobs = []
                for pm in re.finditer(rf"<{tag}\b[^>]*>(.*?)</{tag}>", seg, flags=re.I | re.S):
                    b = re.sub(r"\r\n?", "\n", pm.group(1)).strip()
                    if b:
                        blobs.append(b)
                if blobs:
                    parts[tag] = "\n\n---\n\n".join(blobs)
            return parts

        for i, m in enumerate(obj_starts):
            attr_text = m.group(1)
            # Segment for this Object: from end of start-tag to next Object start (or EOF)
            seg_start = m.end()
            seg_end = obj_starts[i + 1].start() if i + 1 < len(obj_starts) else len(xml_text)
            segment = xml_text[seg_start:seg_end]

            attrs = dict(re.findall(r'(\w+)\s*=\s*"([^"]*)"', attr_text))
            attrs_norm = {k: clean(v) for k, v in attrs.items()}

            fq = attrs_norm.get("fullyQualifiedName", "") or attrs_norm.get("FullyQualifiedName", "") or ""
            nm = attrs_norm.get("name", "") or attrs_norm.get("Name", "") or ""
            object_name = fq if (fq and "." in fq and len(fq) >= max(8, len(nm))) else (nm or fq)

            raw_type_id = attrs_norm.get("type", "") or attrs_norm.get("objectType", "") or attrs_norm.get("ObjectType", "")
            raw_type_name = attrs_norm.get("typeName", "") or attrs_norm.get("objectTypeName", "") or attrs_norm.get("TypeName", "")

            desc = attrs_norm.get("description", "") or attrs_norm.get("Description", "") or ""

            # Collect <Source> tags inside this Object segment (often nested under Parts).
            sources: List[str] = []
            for sm in re.finditer(r"<Source\b[^>]*>(.*?)</Source>", segment, flags=re.I | re.S):
                s = re.sub(r"\r\n?", "\n", sm.group(1)).strip()
                if s:
                    sources.append(s)

            # If no explicit <Source> tags, treat the raw segment text (without tags) as a "source-like" blob.
            if not sources:
                # Remove XML-ish tags; keep plain code/text.
                seg_no_tags = re.sub(r"<[^>]+>", "", segment)
                seg_no_tags = re.sub(r"\r\n?", "\n", seg_no_tags).strip()
                # Only accept if it contains some non-trivial code/text
                if len(seg_no_tags) >= 20:
                    sources.append(seg_no_tags[:200000])  # cap per object

            parts = _extract_part_blobs(segment)
            parts_all = re.sub(r"\r\n?", "\n", segment).strip()
            parts_all = parts_all[:200000]  # cap to keep Excel reasonable

            recs.append({
                "guid": attrs_norm.get("guid", ""),
                "parent_guid": attrs_norm.get("parentGuid", "") or attrs_norm.get("parentGUID", ""),
                "module_guid": attrs_norm.get("moduleGuid", ""),
                "fully_qualified_name": clean(fq),
                "object_name": clean(object_name),
                "raw_type_id": clean(raw_type_id),
                "raw_type_name": clean(raw_type_name),
                "title": "",
                "description": clean(desc),
                "hints": "",
                "sources": sources,
                "parts": parts,
                "parts_all": parts_all,
                "source_file": source_label,
                "text_sample": (segment[:5000]).lower(),
            })
    return recs




def looks_like_object_type_name(name: str) -> bool:
    n = clean(name).lower()
    if not n:
        return False
    # Accept if it matches known aliases or contains strong keywords
    if n in TYPE_ALIASES:
        return True
    strong = ["web", "panel", "transaction", "procedure", "dataprovider", "data provider", "structured", "sdt", "workwith", "rest", "api", "service", "smart"]
    return any(k in n for k in strong)


def extract_type_defs_from_xml(xml_text: str) -> Dict[str, str]:
    """Try to discover type GUID -> type name definitions from XML files.
    This is a best-effort helper for cases where object records only contain type GUID.
    """
    out: Dict[str, str] = {}
    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return out

    for e in root.iter():
        local = e.tag.split("}")[-1].lower()
        guid = clean(e.attrib.get("guid", "") or e.attrib.get("typeGuid", "") or e.attrib.get("typeGUID", ""))
        name = clean(e.attrib.get("name", "") or e.attrib.get("typeName", "") or e.attrib.get("objectTypeName", ""))
        if not name:
            # some formats: <Type guid="...">WebPanel</Type>
            if e.text and len(clean(e.text)) <= 50:
                name = clean(e.text)

        # Heuristics: only accept likely "object type" entries
        if guid and is_uuid(guid) and name and looks_like_object_type_name(name):
            # prefer longer/more descriptive name
            if guid not in out or len(name) > len(out[guid]):
                out[guid] = name

        # Another pattern: element itself is ObjectType with guid+name
        if not guid and local == "objecttype":
            guid = clean(e.attrib.get("id", "") or "")
            if guid and is_uuid(guid) and name and looks_like_object_type_name(name):
                if guid not in out or len(name) > len(out[guid]):
                    out[guid] = name
    return out


# -----------------------------
# Design file iterator
# -----------------------------

def iter_design_files(design_dir: Path) -> List[Tuple[Path, str]]:
    out: List[Tuple[Path, str]] = []
    design_dir = design_dir.resolve()
    tmp_dir = design_dir / ".tmp_extract"
    tmp_dir.mkdir(parents=True, exist_ok=True)

    for p in design_dir.rglob("*"):
        if p.is_dir():
            continue
        suf = p.suffix.lower()
        if suf in (".zip", ".xpz"):
            extract_root = tmp_dir / f"{p.stem}_extracted"
            extract_root.mkdir(parents=True, exist_ok=True)
            try:
                with zipfile.ZipFile(p, "r") as zf:
                    safe_extract_zip(zf, extract_root)
            except Exception:
                continue
            for inner in extract_root.rglob("*"):
                if inner.is_file():
                    label = f"archive:{p.name}:{inner.relative_to(extract_root).as_posix()}"
                    out.append((inner, label))
        else:
            label = str(p.relative_to(design_dir))
            out.append((p, label))
    return out


# -----------------------------
# Build design objects (merge by GUID)
# -----------------------------

def build_design_objects(design_dir: Path, verbose: bool = False) -> List[DesignObject]:
    files = iter_design_files(design_dir)

    # First pass: parse all and collect raw records
    type_defs: Dict[str, str] = {}
    raw_records: List[Dict] = []
    for path, label in files:
        suf = path.suffix.lower()
        try:
            txt = read_text_safely(path)
        except Exception:
            continue

        if suf in (".xml", ".gxobj", ".gxd", ".gxl") or txt.lstrip().startswith("<"):
            type_defs.update(extract_type_defs_from_xml(txt))
            raw_records.extend(extract_objects_from_xml(txt, label))
        elif suf == ".json":
            # JSON: accept list/object
            try:
                data = json.loads(txt)
            except Exception:
                continue
            items = data if isinstance(data, list) else [data]
            for it in items:
                if not isinstance(it, dict):
                    continue
                rec = {
                    "guid": clean(str(it.get("guid", "") or it.get("Guid", ""))),
                    "parent_guid": clean(str(it.get("parentGuid", "") or it.get("ParentGuid", ""))),
                    "module_guid": clean(str(it.get("moduleGuid", "") or it.get("ModuleGuid", ""))),
                    "fully_qualified_name": clean(str(it.get("fullyQualifiedName", "") or it.get("FullyQualifiedName", ""))),
                    "object_name": clean(str(it.get("object_name", "") or it.get("ObjectName", "") or it.get("Name", ""))),
                    "raw_type_id": clean(str(it.get("type", "") or it.get("ObjectType", "") or it.get("Type", ""))),
                    "raw_type_name": clean(str(it.get("typeName", "") or it.get("ObjectTypeName", "") or "")),
                    "title": clean(str(it.get("Title", "") or it.get("Caption", "") or "")),
                    "description": clean(str(it.get("Description", "") or it.get("Desc", "") or "")),
                    "hints": "",
                    "sources": [],
                    "parts": {},
                    "parts_all": "",
                    "source_file": label,
                    "text_sample": txt[:5000].lower(),
                }
                raw_records.append(rec)
        else:
            # plain text fallback: try detect Object header line and Source tags
            # object name candidates
            name = ""
            m = re.search(r"\b(fullyQualifiedName|name)\s*=\s*\"([^\"]+)\"", txt)
            if m:
                name = m.group(2)
            if not name:
                name = path.stem
            # sources in text
            sources = []
            for sm in re.finditer(r"<Source[^>]*>(.*?)</Source>", txt, flags=re.I | re.S):
                s = re.sub(r"\r\n?", "\n", sm.group(1)).strip()
                if s:
                    sources.append(s)
            raw_records.append({
                "guid": "",
                "parent_guid": "",
                "module_guid": "",
                "fully_qualified_name": "",
                "object_name": clean(name),
                "raw_type_id": "",
                "raw_type_name": "",
                "title": "",
                "description": "",
                "hints": "",
                "sources": sources,
                "parts": {"Source": join_sources(sources, limit=30000)} if sources else {},
                "parts_all": "",
                "source_file": label,
                "text_sample": txt[:5000].lower(),
            })

    # Build type registry: type_guid -> type_name when both exist
    type_registry: Dict[str, str] = dict(type_defs)  # seed from type-def XML nodes if any
    for r in raw_records:
        tid = clean(r.get("raw_type_id", ""))
        tname = clean(r.get("raw_type_name", ""))
        if is_uuid(tid) and tname:
            type_registry[tid] = tname
        elif is_uuid(tid) and tid and tid not in type_registry:
            # 末尾4桁の固定マッピングは世代差でズレる可能性があるため採用しない。
            # Folder/Category（…0006）のみは画像指針に従い、必要なら後段の Part 判定で拾う。
            pass

    # Merge records by guid (preferred), else by fully qualified name, else by object_name+raw_type_id
    merged: Dict[str, Dict] = {}
    def key_of(r: Dict) -> str:
        guid = clean(r.get("guid", ""))
        if guid:
            return f"guid:{guid}"
        fq = clean(r.get("fully_qualified_name", ""))
        if fq:
            return f"fq:{fq}"
        name = clean(r.get("object_name", ""))
        tid = clean(r.get("raw_type_id", ""))
        return f"name:{name}|type:{tid}"

    def score(r: Dict) -> int:
        return len(clean(r.get("object_name", ""))) + len(clean(r.get("description", ""))) + len(clean(r.get("title", ""))) + len(clean(r.get("raw_type_name", "")))

    for r in raw_records:
        k = key_of(r)
        if k not in merged:
            merged[k] = r
        else:
            # merge: keep higher score metadata, always append sources, merge hints
            cur = merged[k]
            if score(r) > score(cur):
                # keep sources/hints from cur too
                r_sources = list(cur.get("sources", [])) + list(r.get("sources", []))
                r_hints = ",".join(sorted(set(filter(None, (cur.get("hints","") + "," + r.get("hints","")).split(",")))))
                r["sources"] = r_sources
                r["hints"] = r_hints.strip(",")
                # merge parts
                r["parts"] = merge_parts_dict(cur.get("parts"), r.get("parts"))
                r["parts_all"] = _join_blobs([cur.get("parts_all",""), r.get("parts_all","")], limit=40000)
                merged[k] = r
            else:
                cur["sources"] = list(cur.get("sources", [])) + list(r.get("sources", []))
                cur["hints"] = ",".join(sorted(set(filter(None, (cur.get("hints","") + "," + r.get("hints","")).split(","))))).strip(",")

                # merge parts
                cur["parts"] = merge_parts_dict(cur.get("parts"), r.get("parts"))
                cur["parts_all"] = _join_blobs([cur.get("parts_all",""), r.get("parts_all","")], limit=40000)
    objs: List[DesignObject] = []
    for r in merged.values():
        obj_name = clean(r.get("object_name", "")) or "UnknownObject"
        fq = clean(r.get("fully_qualified_name", ""))
        guid = clean(r.get("guid", ""))
        parent_guid = clean(r.get("parent_guid", ""))
        module_guid = clean(r.get("module_guid", ""))
        raw_type_id = clean(r.get("raw_type_id", ""))
        raw_type_name = clean(r.get("raw_type_name", ""))

        # Fill missing raw_type_name from registry if possible
        if not raw_type_name and is_uuid(raw_type_id) and raw_type_id in type_registry:
            raw_type_name = type_registry[raw_type_id]
        # Determine final object_type
        # 1) まず raw_type_name（明示）を正規化
        base_type = "Unknown"
        if raw_type_name:
            base_type = normalize_object_type(raw_type_name, filename=r.get("source_file",""), object_name=obj_name)
        else:
            # if raw_type_id is not uuid and looks like type name, normalize directly
            if raw_type_id and not is_uuid(raw_type_id):
                base_type = normalize_object_type(raw_type_id, filename=r.get("source_file",""), object_name=obj_name)
            else:
                guessed = guess_type_from_text(r.get("text_sample",""), obj_name, r.get("source_file",""))
                base_type = normalize_object_type(guessed, filename=r.get("source_file",""), object_name=obj_name)

        # 2) 画像の「Part構成で判定」を適用して修正（特に GUID だけで型が不明なケースや WWP 判定）
        parts_norm = _canonicalize_parts(r.get("parts") or {})
        r["parts"] = parts_norm  # 出力列の安定化（Source vs source など）
        type_by_parts = infer_object_type_from_parts(parts_norm, r.get("parts_all",""), r.get("text_sample",""), raw_type_id)

        obj_type = base_type
        if type_by_parts:
            # Folder は最優先
            if type_by_parts == "Folder":
                obj_type = "Folder"
            # WWP は raw が Web Panel でも上書き
            elif base_type == "WP" and type_by_parts == "WWP":
                obj_type = "WWP"
            # raw が Unknown/弱い場合は Parts 判定を採用
            elif base_type in ("Unknown", ""):
                obj_type = type_by_parts

        src_text = join_sources(list(dict.fromkeys([s for s in r.get("sources", []) if s])), limit=10000)

        o = DesignObject(
            object_name=obj_name,
            object_type=obj_type,
            source_file=r.get("source_file", ""),
            guid=guid,
            parent_guid=parent_guid,
            module_guid=module_guid,
            fully_qualified_name=fq,
            raw_type_id=raw_type_id,
            raw_type_name=raw_type_name,
            source=src_text,
            parts=r.get("parts", {}) or {},
            parts_all=clean(r.get("parts_all","")),
            title=clean(r.get("title", "")),
            description=clean(r.get("description", "")),
            raw_hints=clean(r.get("hints", "")),
        )
        # Infer entity/actions/features
        o.inferred_entity = infer_entity(o)
        o.inferred_actions = infer_actions(o)
        o.feature_names = ",".join(make_feature_names(o))
        objs.append(o)

    if verbose:
        print(f"[INFO] Parsed records: {len(raw_records)} | Unique objects: {len(objs)}", file=sys.stderr)
        # type distribution
        dist: Dict[str, int] = {}
        for o in objs:
            dist[o.object_type] = dist.get(o.object_type, 0) + 1
        print(f"[INFO] Type distribution: {dist}", file=sys.stderr)

    return sorted(objs, key=lambda x: (x.object_type, x.object_name))


# -----------------------------
# Feature inference (same spirit as v3, slightly refined)
# -----------------------------

def infer_entity(o: DesignObject) -> str:
    name = o.object_name
    # Remove namespace-like prefix in fully qualified names: A.B.C -> take last segment as entity base
    base = name.split(".")[-1] if "." in name else name
    # Strip common suffixes
    base = re.sub(r"(WW|WWP|View|Detail|Insert|Update|Delete|Prompt|Picker|List|Grid|Panel|SD|DP|Proc|Trn)\d*$", "", base, flags=re.I)
    return base or name


ACTION_PATTERNS = [
    (r"\b(list|grid|for\s+each)\b", "一覧"),
    (r"\b(search|filter|query)\b", "検索"),
    (r"\b(insert|create|new)\b", "新規"),
    (r"\b(update|edit|modify)\b", "更新"),
    (r"\b(delete|remove)\b", "削除"),
    (r"\b(export|print|report)\b", "帳票/出力"),
    (r"\b(import)\b", "取込"),
    (r"\b(sync)\b", "同期"),
    (r"\b(login|auth)\b", "認証"),
]


def infer_actions(o: DesignObject) -> str:
    t = o.object_type
    text = " ".join([o.object_name, o.title, o.description, o.raw_hints]).lower()

    actions: List[str] = []
    # Type-specific defaults
    if t in ("Transaction",):
        actions = ["新規", "更新", "削除"]
    elif t in ("DP", "Procedure"):
        # prefer description/title keywords for DP/Procedure
        if any(k in (o.description + " " + o.title) for k in ["発行", "帳票", "印刷", "明細", "一覧"]):
            actions.append("帳票/出力")
    elif t in ("WWP", "WP", "SD"):
        # UI panels often list+search
        if "grid" in text or "for each" in text or "一覧" in (o.title + o.description):
            actions.append("一覧")
        if "search" in text or "filter" in text or "検索" in (o.title + o.description):
            actions.append("検索")
    elif t == "SDT":
        actions.append("データモデル")

    # Pattern scan fallback
    for pat, act in ACTION_PATTERNS:
        if re.search(pat, text, flags=re.I):
            if act not in actions:
                actions.append(act)

    if not actions:
        actions = ["機能"]

    return ",".join(actions)


def make_feature_names(o: DesignObject) -> List[str]:
    entity = o.inferred_entity or o.object_name
    acts = [a.strip() for a in (o.inferred_actions or "").split(",") if a.strip()]
    # If DP/Proc has a good description, use it as feature name (works well for “発行DP”等)
    if o.object_type in ("DP", "Procedure") and o.description:
        return [o.description]

    if "一覧" in acts and "検索" in acts:
        return [f"{entity} 一覧検索"]
    if o.object_type == "Transaction" and all(a in acts for a in ("新規", "更新", "削除")):
        return [f"{entity} メンテナンス（新規/更新/削除）"]

    feats = []
    for a in acts[:3]:
        feats.append(f"{entity} {a}")
    return feats or [f"{entity} 機能"]


# -----------------------------
# Output tables
# -----------------------------

def build_feature_design_rows(objs: List[DesignObject]) -> List[Dict]:
    rows: List[Dict] = []
    for o in objs:
        if o.object_type == "Folder":
            continue
        feats = [x.strip() for x in o.feature_names.split(",") if x.strip()] or [f"{o.inferred_entity or o.object_name} 機能"]
        for f in feats:
            rows.append({
                "機能名": f,
                "設計オブジェクト種別": o.object_type,
                "設計オブジェクト名": o.object_name,
                "FQ名": o.fully_qualified_name,
                "Object GUID": o.guid,
                "Parent GUID": o.parent_guid,
                "Module GUID": o.module_guid,
                "Raw Type ID": o.raw_type_id,
                "Raw Type Name": o.raw_type_name,
                "エンティティ(推定)": o.inferred_entity,
                "アクション(推定)": o.inferred_actions,
                "タイトル/Caption": o.title,
                "説明/Doc": o.description,
                "Source(抽出)": o.source,
                # Parts (ユーザー記述 Part)
                "Part_Source": o.parts.get("Source","") if "Source" in TYPE_TO_WRITABLE_PARTS.get(o.object_type, PART_TAGS_ORDER) else "",
                "Part_Rules": o.parts.get("Rules","") if "Rules" in TYPE_TO_WRITABLE_PARTS.get(o.object_type, PART_TAGS_ORDER) else "",
                "Part_Variables": o.parts.get("Variables","") if "Variables" in TYPE_TO_WRITABLE_PARTS.get(o.object_type, PART_TAGS_ORDER) else "",
                "Part_Parm": o.parts.get("Parm","") if "Parm" in TYPE_TO_WRITABLE_PARTS.get(o.object_type, PART_TAGS_ORDER) else "",
                "Part_Events": o.parts.get("Events","") if "Events" in TYPE_TO_WRITABLE_PARTS.get(o.object_type, PART_TAGS_ORDER) else "",
                "Part_Layout": o.parts.get("Layout","") if "Layout" in TYPE_TO_WRITABLE_PARTS.get(o.object_type, PART_TAGS_ORDER) else "",
                "Part_Attributes": o.parts.get("Attributes","") if "Attributes" in TYPE_TO_WRITABLE_PARTS.get(o.object_type, PART_TAGS_ORDER) else "",
                "Part_Structure": o.parts.get("Structure","") if "Structure" in TYPE_TO_WRITABLE_PARTS.get(o.object_type, PART_TAGS_ORDER) else "",
                "Part_Style": o.parts.get("Style","") if "Style" in TYPE_TO_WRITABLE_PARTS.get(o.object_type, PART_TAGS_ORDER) else "",
                "Parts_All_Raw": o.parts_all,
                "Hints": o.raw_hints,
                "元ファイル": o.source_file,
            })
    return rows


def build_feature_group_rows(feature_design_rows: List[Dict]) -> List[Dict]:
    by_feature: Dict[str, List[Dict]] = {}
    for r in feature_design_rows:
        by_feature.setdefault(r["機能名"], []).append(r)

    out: List[Dict] = []
    for feat, rows in sorted(by_feature.items(), key=lambda x: x[0]):
        objs = sorted(set(f'{r["設計オブジェクト種別"]}:{r["設計オブジェクト名"]}' for r in rows))
        out.append({
            "機能名": feat,
            "対象オブジェクト数": len(objs),
            "設計オブジェクト一覧": "\n".join(objs),
        })
    return out


# -----------------------------
# Java mapping (same as v3)
# -----------------------------

def iter_java_files(java_dir: Path) -> List[Path]:
    return [p for p in java_dir.rglob("*.java") if p.is_file()]


def normalize_name_for_match(name: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "", name or "").lower()


def split_camel_tokens(s: str) -> List[str]:
    s = re.sub(r"[^a-zA-Z0-9]+", " ", s)
    parts = re.findall(r"[A-Z]?[a-z]+|[A-Z]+(?![a-z])|\d+", s)
    return [p.lower() for p in parts if p]


def index_java_files(java_dir: Path) -> Tuple[Dict[str, List[Path]], Dict[str, List[Path]]]:
    by_stem: Dict[str, List[Path]] = {}
    by_token: Dict[str, List[Path]] = {}
    for p in iter_java_files(java_dir):
        stem = normalize_name_for_match(p.stem)
        by_stem.setdefault(stem, []).append(p)
        for tok in split_camel_tokens(p.stem):
            by_token.setdefault(tok, []).append(p)
    return by_stem, by_token


def match_java_for_object(obj_name: str, by_stem: Dict[str, List[Path]], by_token: Dict[str, List[Path]]) -> List[Dict]:
    name_norm = normalize_name_for_match(obj_name.split(".")[-1])  # use last segment for code gen
    candidates: List[Tuple[Path, str, float]] = []

    # Strong: stem match
    if name_norm in by_stem:
        for p in by_stem[name_norm]:
            candidates.append((p, "filename_stem_match", 0.95))

    # Medium: token pool + inspect head
    if not candidates:
        toks = split_camel_tokens(obj_name.split(".")[-1])
        pool: List[Path] = []
        if toks and toks[0] in by_token:
            pool = by_token[toks[0]][:2000]
        else:
            pool = []
        for p in pool:
            try:
                txt = read_text_safely(p, max_bytes=200_000)
            except Exception:
                continue
            t = txt.lower()
            if re.search(r"\bclass\s+" + re.escape(obj_name.split(".")[-1]) + r"\b", txt):
                candidates.append((p, "class_name_match_in_file", 0.85))
            elif name_norm and name_norm in normalize_name_for_match(t):
                candidates.append((p, "content_contains_object_name", 0.60))

    # Weak: stem contains
    if not candidates and name_norm:
        for stem, paths in by_stem.items():
            if name_norm in stem:
                for p in paths[:10]:
                    candidates.append((p, "stem_contains_name", 0.40))
                break

    # Dedup keep best score per file
    best: Dict[str, Tuple[Path, str, float]] = {}
    for p, method, conf in candidates:
        k = str(p)
        if k not in best or conf > best[k][2]:
            best[k] = (p, method, conf)

    ranked = sorted(best.values(), key=lambda x: (-x[2], str(x[0])))
    out = [{
        "設計オブジェクト名": obj_name,
        "Javaファイル": str(p),
        "マッチ方式": method,
        "信頼度": conf,
    } for p, method, conf in ranked[:5]]

    if not out:
        out = [{
            "設計オブジェクト名": obj_name,
            "Javaファイル": "",
            "マッチ方式": "not_found",
            "信頼度": 0.0,
        }]
    return out


def build_design_java_rows(objs: List[DesignObject], java_dir: Path) -> List[Dict]:
    by_stem, by_token = index_java_files(java_dir)
    rows: List[Dict] = []
    for o in objs:
        if o.object_type == "Theme":
            continue
        if o.object_type == "Folder":
            continue
        rows.extend(match_java_for_object(o.object_name, by_stem, by_token))
    return rows


# -----------------------------
# Excel writer
# -----------------------------

def autosize_worksheet(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
            except Exception:
                v = ""
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(60, max(10, int(max_len * 0.9) + 2))


def write_sheet(wb, name: str, rows: List[Dict]) -> None:
    ws = wb.create_sheet(title=name)
    if not rows:
        ws.append(["(no data)"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = wrap
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)


def save_excel(out_path: Path, feature_design: List[Dict], feature_group: List[Dict], design_java: List[Dict]) -> None:
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    write_sheet(wb, "Feature_Design", feature_design)
    write_sheet(wb, "Feature_Group", feature_group)
    write_sheet(wb, "Design_Java", design_java)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


# -----------------------------
# Main
# -----------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="GeneXus traceability builder v9 (infer ObjectType by Parts composition)")
    ap.add_argument("--design_dir", required=True, help="Design export directory (may include XPZ/ZIP)")
    ap.add_argument("--java_dir", default="", help="Generated Java source root (optional)")
    ap.add_argument("--out", required=True, help="Output xlsx path")
    ap.add_argument("--dump_objects_json", default="", help="Dump parsed objects to JSON for debugging (optional)")
    ap.add_argument("-v", "--verbose", action="store_true", help="Verbose logging")
    args = ap.parse_args()

    design_dir = Path(args.design_dir)
    java_dir = Path(args.java_dir) if args.java_dir else None
    out_path = Path(args.out)

    objs = build_design_objects(design_dir, verbose=args.verbose)

    if args.dump_objects_json:
        dump_path = Path(args.dump_objects_json)
        dump_path.parent.mkdir(parents=True, exist_ok=True)
        with open(dump_path, "w", encoding="utf-8") as f:
            json.dump([o.__dict__ for o in objs], f, ensure_ascii=False, indent=2)

    feature_design = build_feature_design_rows(objs)
    feature_group = build_feature_group_rows(feature_design)

    design_java: List[Dict] = []
    if java_dir and java_dir.exists():
        design_java = build_design_java_rows(objs, java_dir)

    save_excel(out_path, feature_design, feature_group, design_java)

    if args.verbose:
        print(f"[INFO] Objects: {len(objs)} | Feature rows: {len(feature_design)} | Out: {out_path}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
