#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gx_crud_table_generator.py

One-shot GeneXus CRUD matrix generator.

Purpose
-------
Given a GeneXus system export directory (XML/TXT/JSON/XPZ/ZIP) and optionally
its generated Java source directory, generate CRUD tables for all inferred
business functions.

Outputs
-------
- Excel workbook (.xlsx): Function CRUD matrix, Object CRUD detail, Entity summary,
  Design-Java traceability, Parse log.
- JSON file: machine-readable full result.
- Markdown file: compact human-readable report.

Optional Claude support
-----------------------
The default parser is deterministic and does not require Claude. If the project
export is sparse or naming is unusual, use --claude auto/api/cli to let Claude
refine ambiguous object/function/action inference. The script will skip Claude
cleanly when no credentials/CLI are available.

Typical usage
-------------
python gx_crud_table_generator.py \
  --design-dir /path/to/gx/export \
  --java-dir /path/to/JavaModel/src/main/java \
  --out-xlsx ./gx_crud_matrix.xlsx \
  --out-json ./gx_crud_matrix.json \
  --out-md ./gx_crud_matrix.md

Dependencies
------------
Python 3.9+. Excel output needs openpyxl. JSON/Markdown work without openpyxl.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import shutil
import subprocess
import sys
import textwrap
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

TEXT_EXTENSIONS = {
    ".xml", ".json", ".txt", ".gxd", ".gxobj", ".gxl", ".gxw",
    ".properties", ".ini", ".yaml", ".yml", ".java", ".cs", ".sql",
    ".xpz", ".zip",
}
BINARY_SKIP_EXTENSIONS = {
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".ico", ".pdf", ".docx",
    ".xlsx", ".xls", ".pptx", ".class", ".jar", ".war", ".ear", ".exe",
    ".dll", ".so", ".dylib", ".ttf", ".woff", ".woff2",
}
ARCHIVE_EXTENSIONS = {".zip", ".xpz"}
ENCODINGS = ("utf-8", "utf-8-sig", "cp932", "shift_jis", "gbk", "latin-1")
CRUD_ACTIONS = ("C", "R", "U", "D")
ACTION_LABEL = {"C": "Create", "R": "Read", "U": "Update", "D": "Delete"}

# GeneXus object type signals. Keep the list permissive because exports vary.
TYPE_KEYWORDS: Dict[str, Sequence[str]] = {
    "Transaction": ("transaction", "trn", "business component"),
    "WWP": ("workwith", "work with", "workwithplus", "wwp", "work with plus"),
    "WP": ("web panel", "webpanel", "wp"),
    "SD": ("smart device", "sd panel", "sdpanel", "mobile panel"),
    "SDT": ("structured data type", "sdt"),
    "Procedure": ("procedure", "proc"),
    "DP": ("data provider", "dataprovider", "dp"),
    "REST": ("rest", "web service", "webservice", "expose as web service"),
    "API": ("api object", "apiobject", "openapi", "swagger"),
    "Query": ("query object", "gx query"),
    "Dashboard": ("dashboard"),
    "Theme": ("theme"),
}

CRUD_PATTERNS: Dict[str, List[Tuple[re.Pattern[str], float, str]]] = {
    "C": [
        (re.compile(r"\b(insert|create|add|new)\b", re.I), 0.70, "create-like keyword"),
        (re.compile(r"\.(insert|save)\s*\(", re.I), 0.88, "business component insert/save call"),
        (re.compile(r"\binsert\s+into\b", re.I), 0.95, "SQL INSERT"),
        (re.compile(r"\bmethod\s*[:=]\s*post\b|\bhttp\s*post\b|\bPOST\s+/", re.I), 0.85, "HTTP POST"),
    ],
    "R": [
        (re.compile(r"\b(for\s+each|grid|load|read|view|detail|search|filter|find|query|browse|list|prompt|select)\b", re.I), 0.72, "read/list/search keyword"),
        (re.compile(r"\bselect\b.+\bfrom\b", re.I | re.S), 0.92, "SQL SELECT"),
        (re.compile(r"\bwhere\b|\border\s+by\b", re.I), 0.70, "query condition"),
        (re.compile(r"\bmethod\s*[:=]\s*get\b|\bhttp\s*get\b|\bGET\s+/", re.I), 0.85, "HTTP GET"),
    ],
    "U": [
        (re.compile(r"\b(update|edit|modify|change|set)\b", re.I), 0.70, "update-like keyword"),
        (re.compile(r"\.(update)\s*\(", re.I), 0.88, "business component update call"),
        (re.compile(r"\bupdate\s+[A-Za-z_][A-Za-z0-9_]*\s+set\b", re.I), 0.95, "SQL UPDATE"),
        (re.compile(r"\bmethod\s*[:=]\s*(put|patch)\b|\bhttp\s*(put|patch)\b|\b(PUT|PATCH)\s+/", re.I), 0.85, "HTTP PUT/PATCH"),
    ],
    "D": [
        (re.compile(r"\b(delete|remove)\b", re.I), 0.70, "delete-like keyword"),
        (re.compile(r"\.(delete)\s*\(", re.I), 0.88, "business component delete call"),
        (re.compile(r"\bdelete\s+from\b", re.I), 0.95, "SQL DELETE"),
        (re.compile(r"\bmethod\s*[:=]\s*delete\b|\bhttp\s*delete\b|\bDELETE\s+/", re.I), 0.85, "HTTP DELETE"),
    ],
}

WWP_SUFFIX_RULES: List[Tuple[re.Pattern[str], Sequence[str], str]] = [
    (re.compile(r"(WW|WorkWith|List|Grid|Browse)$", re.I), ("R",), "WWP/list suffix"),
    (re.compile(r"(View|Detail|Display)$", re.I), ("R",), "detail/view suffix"),
    (re.compile(r"(Prompt|Select|Picker|Lookup)$", re.I), ("R",), "prompt/select suffix"),
    (re.compile(r"(Insert|Create|New|Add)$", re.I), ("C",), "insert/create suffix"),
    (re.compile(r"(Update|Edit|Modify)$", re.I), ("R", "U"), "update/edit suffix"),
    (re.compile(r"(Delete|Remove)$", re.I), ("D",), "delete/remove suffix"),
]

OBJECT_SUFFIXES = (
    "WorkWith", "WW", "List", "Grid", "Browse", "View", "Detail", "Display",
    "Prompt", "Select", "Picker", "Lookup", "Insert", "Create", "New", "Add",
    "Update", "Edit", "Modify", "Delete", "Remove", "Panel", "WP", "SD",
    "Proc", "Procedure", "DataProvider", "DP", "BC",
)

IGNORE_DIR_PARTS = {
    ".git", ".svn", ".hg", "node_modules", "target", "bin", "obj", "build",
    ".tmp_extract", "__pycache__",
}


@dataclass
class SourceText:
    path_label: str
    suffix: str
    text: str
    size_bytes: int


@dataclass
class Evidence:
    action: str
    weight: float
    reason: str
    source: str
    snippet: str = ""


@dataclass
class DesignObject:
    object_name: str
    object_type: str
    source_file: str
    title: str = ""
    description: str = ""
    entity: str = ""
    entities: List[str] = field(default_factory=list)
    actions: Dict[str, float] = field(default_factory=lambda: {a: 0.0 for a in CRUD_ACTIONS})
    evidence: List[Evidence] = field(default_factory=list)
    feature_names: List[str] = field(default_factory=list)
    raw_hints: str = ""
    java_files: List[str] = field(default_factory=list)
    java_tables: List[str] = field(default_factory=list)


@dataclass
class JavaFileInfo:
    path: Path
    rel_path: str
    stem_norm: str
    class_names: List[str]
    tokens: List[str]
    actions: Dict[str, float]
    evidence: List[Evidence]
    tables: List[str]
    text_sample: str = ""


@dataclass
class FunctionAggregate:
    function_id: str
    function_name: str
    primary_entity: str
    entities: List[str]
    actions: Dict[str, float]
    design_objects: List[str]
    object_types: List[str]
    java_files: List[str]
    evidence: List[Evidence]


def read_text_from_bytes(data: bytes, max_bytes: int) -> str:
    data = data[:max_bytes]
    for enc in ENCODINGS:
        try:
            return data.decode(enc)
        except Exception:
            pass
    return data.decode("utf-8", errors="ignore")


def read_text_safely(path: Path, max_bytes: int = 2_000_000) -> str:
    return read_text_from_bytes(path.read_bytes(), max_bytes=max_bytes)


def normalize_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def is_probable_name(s: str) -> bool:
    if not s:
        return False
    s = str(s).strip()
    if len(s) > 120:
        return False
    if s.lower() in {"name", "type", "title", "caption", "true", "false", "none", "object"}:
        return False
    return bool(re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", s))


def camel_tokens(name: str) -> List[str]:
    parts = re.findall(r"[A-Z]+(?=[A-Z][a-z])|[A-Z]?[a-z]+|[0-9]+", name or "")
    return [p.lower() for p in parts if p]


def unique_keep_order(values: Iterable[str]) -> List[str]:
    seen: Set[str] = set()
    out: List[str] = []
    for v in values:
        v = str(v).strip()
        if not v:
            continue
        key = v.lower()
        if key not in seen:
            out.append(v)
            seen.add(key)
    return out


def compact_snippet(text: str, match: Optional[re.Match[str]] = None, width: int = 160) -> str:
    text = re.sub(r"\s+", " ", text or "").strip()
    if not text:
        return ""
    if match:
        start = max(0, match.start() - width // 3)
        end = min(len(text), match.end() + width // 3)
        return text[start:end][:width]
    return text[:width]


def safe_rel(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except Exception:
        return str(path)


def iter_plain_files(root: Path, max_file_bytes: int) -> Iterable[SourceText]:
    for p in sorted(root.rglob("*")):
        if p.is_dir():
            continue
        if any(part in IGNORE_DIR_PARTS for part in p.parts):
            continue
        suffix = p.suffix.lower()
        if suffix in BINARY_SKIP_EXTENSIONS:
            continue
        if suffix not in TEXT_EXTENSIONS and suffix not in ARCHIVE_EXTENSIONS:
            # GeneXus exports can be extensionless, but avoid huge unknown binaries.
            if p.stat().st_size > max_file_bytes:
                continue
        if suffix in ARCHIVE_EXTENSIONS:
            yield from iter_archive_members(p, p.name, max_file_bytes=max_file_bytes)
        else:
            try:
                text = read_text_safely(p, max_bytes=max_file_bytes)
                yield SourceText(safe_rel(p, root), suffix, text, p.stat().st_size)
            except Exception as exc:
                yield SourceText(safe_rel(p, root), suffix, f"", 0)


def iter_archive_members(path: Path, label: str, max_file_bytes: int, depth: int = 0) -> Iterable[SourceText]:
    if depth > 2:
        return
    try:
        data = path.read_bytes()
        zf = zipfile.ZipFile(io.BytesIO(data))
    except Exception:
        return
    with zf:
        for info in sorted(zf.infolist(), key=lambda x: x.filename):
            if info.is_dir():
                continue
            member = info.filename
            if any(part in IGNORE_DIR_PARTS for part in Path(member).parts):
                continue
            suffix = Path(member).suffix.lower()
            if suffix in BINARY_SKIP_EXTENSIONS:
                continue
            source_label = f"archive:{label}:{member}"
            try:
                raw = zf.read(info, pwd=None)
            except Exception:
                continue
            if suffix in ARCHIVE_EXTENSIONS:
                tmp_name = Path(member).name
                try:
                    nested = zipfile.ZipFile(io.BytesIO(raw))
                    with nested:
                        for ninfo in sorted(nested.infolist(), key=lambda x: x.filename):
                            if ninfo.is_dir():
                                continue
                            nsuffix = Path(ninfo.filename).suffix.lower()
                            if nsuffix in BINARY_SKIP_EXTENSIONS:
                                continue
                            nraw = nested.read(ninfo)
                            text = read_text_from_bytes(nraw, max_bytes=max_file_bytes)
                            yield SourceText(f"archive:{label}:{member}:{ninfo.filename}", nsuffix, text, len(nraw))
                except Exception:
                    continue
            else:
                text = read_text_from_bytes(raw, max_bytes=max_file_bytes)
                yield SourceText(source_label, suffix, text, len(raw))


def pick_first(patterns: Sequence[str], text: str) -> str:
    for pat in patterns:
        m = re.search(pat, text, flags=re.I | re.S | re.M)
        if m:
            val = re.sub(r"\s+", " ", str(m.group(1))).strip()
            if val:
                return val
    return ""


def guess_type_from_text(text: str, filename: str) -> str:
    t = (text[:50000] + "\n" + filename).lower()
    best_type, best_score = "Unknown", 0
    for object_type, keywords in TYPE_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in t)
        if score > best_score:
            best_type, best_score = object_type, score
    return best_type if best_score else "Unknown"


def normalize_object_type(raw_type: str, guessed: str, filename: str, text: str) -> str:
    rt = (raw_type or "").strip().lower()
    for normalized, keywords in TYPE_KEYWORDS.items():
        for kw in keywords:
            if rt == kw or kw in rt:
                return normalized
    if guessed != "Unknown":
        return guessed
    fn = filename.lower()
    if "workwith" in fn or fn.endswith(("ww.xml", "wwp.xml")):
        return "WWP"
    if "transaction" in fn or "trn" in fn:
        return "Transaction"
    if "procedure" in fn or "proc" in fn:
        return "Procedure"
    if "dataprovider" in fn or re.search(r"(^|[_\-.])dp([_\-.]|$)", fn):
        return "DP"
    if "sdt" in fn:
        return "SDT"
    if "webpanel" in fn or re.search(r"(^|[_\-.])wp([_\-.]|$)", fn):
        return "WP"
    return "Unknown"


def extract_json_objects(data: Any, source_label: str) -> List[Tuple[str, str, str, str, str]]:
    """Return tuples: name, type, title, desc, hints."""
    out: List[Tuple[str, str, str, str, str]] = []
    if isinstance(data, dict):
        name = data.get("ObjectName") or data.get("ObjName") or data.get("Name") or data.get("name")
        otype = data.get("ObjectType") or data.get("Type") or data.get("type") or data.get("Category")
        if name and is_probable_name(str(name)):
            title = data.get("Title") or data.get("Caption") or data.get("FormCaption") or ""
            desc = data.get("Description") or data.get("Documentation") or data.get("Doc") or ""
            hints = ",".join(k for k in ("Rules", "Events", "Conditions", "Source") if k in data)
            out.append((str(name), str(otype or ""), str(title or ""), str(desc or ""), hints))
        for v in data.values():
            out.extend(extract_json_objects(v, source_label))
    elif isinstance(data, list):
        for item in data:
            out.extend(extract_json_objects(item, source_label))
    return out


def extract_name_type_title(src: SourceText) -> List[Tuple[str, str, str, str, str]]:
    text = src.text
    filename = Path(src.path_label).name
    suffix = src.suffix.lower()

    if suffix == ".json":
        try:
            data = json.loads(text)
            extracted = extract_json_objects(data, src.path_label)
            if extracted:
                return extracted
        except Exception:
            pass

    # XML-ish or text fallback. GeneXus exports vary a lot; use broad patterns.
    name = pick_first([
        r"<\s*(?:ObjectName|ObjName|QualifiedName|Name)\s*>\s*([^<\s][^<]{0,120}?)\s*</\s*(?:ObjectName|ObjName|QualifiedName|Name)\s*>",
        r"(?:ObjectName|ObjName|Name)\s*[:=]\s*([A-Za-z_][A-Za-z0-9_]*)",
        r'"\s*(?:ObjectName|ObjName|Name|name)\s*"\s*:\s*"\s*([A-Za-z_][A-Za-z0-9_]*)\s*"',
        r"\b(?:Web\s*Panel|Transaction|Procedure|Data\s*Provider|SDT|Smart\s*Device)\s*[:=]\s*([A-Za-z_][A-Za-z0-9_]*)",
    ], text)

    if not is_probable_name(name):
        # Some XML files contain many <Name> nodes; prefer file stem when clean.
        stem = Path(filename).stem
        if is_probable_name(stem):
            name = stem
        else:
            candidates = re.findall(r"<\s*Name\s*>\s*([A-Za-z_][A-Za-z0-9_]{1,80})\s*</\s*Name\s*>", text, flags=re.I)
            name = next((c for c in candidates if is_probable_name(c)), "")

    raw_type = pick_first([
        r"<\s*(?:ObjectType|Type|Category|Class|ObjCls)\s*>\s*([^<]{1,100}?)\s*</\s*(?:ObjectType|Type|Category|Class|ObjCls)\s*>",
        r"(?:ObjectType|Type|Category|Class)\s*[:=]\s*([^\r\n]{1,100})",
        r'"\s*(?:ObjectType|Type|Category|Class|type)\s*"\s*:\s*"\s*([^\"]{1,100})\s*"',
    ], text)
    title = pick_first([
        r"<\s*(?:Title|Caption|FormCaption|Text)\s*>\s*([^<]{1,250}?)\s*</\s*(?:Title|Caption|FormCaption|Text)\s*>",
        r"(?:Title|Caption|FormCaption)\s*[:=]\s*([^\r\n]{1,250})",
    ], text)
    desc = pick_first([
        r"<\s*(?:Description|Desc|Documentation|Comment)\s*>\s*([^<]{1,500}?)\s*</\s*(?:Description|Desc|Documentation|Comment)\s*>",
        r"(?:Description|Documentation|Doc)\s*[:=]\s*([^\r\n]{1,500})",
    ], text)
    keyword_hits = re.findall(r"\b(Search|Filter|Grid|For\s+each|Link|Call|Export|Import|Approve|Login|Sync|REST|DataProvider|Transaction|Insert|Update|Delete)\b", text, flags=re.I)
    hints = ",".join(sorted(set(k.lower() for k in keyword_hits)))[:500]

    if name and is_probable_name(name):
        return [(name, raw_type, title, desc, hints)]
    return []


def strip_object_suffix(name: str) -> str:
    if not name:
        return ""
    n = name
    # WorkWithCustomer -> Customer
    m = re.match(r"WorkWith(.+)$", n, flags=re.I)
    if m and is_probable_name(m.group(1)):
        n = m.group(1)
    # DeleteCustomer / GetCustomer / UpdateCustomer -> Customer
    for prefix in ("Create", "Insert", "Add", "New", "Get", "Find", "Query", "Search", "List", "Load", "View", "Update", "Edit", "Modify", "Delete", "Remove"):
        if n.lower().startswith(prefix.lower()) and len(n) > len(prefix) + 1:
            rest = n[len(prefix):]
            if is_probable_name(rest):
                n = rest
                break
    for suffix in sorted(OBJECT_SUFFIXES, key=len, reverse=True):
        if n.lower().endswith(suffix.lower()) and len(n) > len(suffix) + 1:
            return n[: -len(suffix)]
    return n


def infer_entities(object_name: str, object_type: str, text: str) -> List[str]:
    entities: List[str] = []
    patterns = [
        r"\b(?:BaseTrn|BaseTransaction|Transaction|Trn)\s*[:=]\s*([A-Za-z_][A-Za-z0-9_]*)",
        r"<\s*(?:BaseTrn|BaseTransaction|TransactionName|TableName)\s*>\s*([A-Za-z_][A-Za-z0-9_]*)\s*</",
        r"\b(?:from|into|update|delete\s+from)\s+([A-Za-z_][A-Za-z0-9_.$]*)",
    ]
    for pat in patterns:
        for m in re.finditer(pat, text, flags=re.I):
            token = m.group(1).split(".")[-1]
            if is_probable_name(token):
                entities.append(token)
    if object_type == "Transaction" and is_probable_name(object_name):
        entities.insert(0, object_name)
    stripped = strip_object_suffix(object_name)
    if is_probable_name(stripped):
        entities.append(stripped)
    return unique_keep_order(entities) or ([object_name] if object_name else [])


def add_evidence(obj: DesignObject, action: str, weight: float, reason: str, source: str, snippet: str = "") -> None:
    if action not in CRUD_ACTIONS:
        return
    obj.actions[action] = max(obj.actions.get(action, 0.0), float(weight))
    obj.evidence.append(Evidence(action=action, weight=float(weight), reason=reason, source=source, snippet=snippet[:220]))


def infer_actions_from_text(obj: DesignObject, text: str) -> None:
    searchable = text[:500000]
    for action, patterns in CRUD_PATTERNS.items():
        for rgx, weight, reason in patterns:
            m = rgx.search(searchable)
            if m:
                add_evidence(obj, action, weight, reason, obj.source_file, compact_snippet(searchable, m))

    # Object-type defaults, useful when design export does not include body/rules.
    if obj.object_type == "Transaction":
        for action in CRUD_ACTIONS:
            add_evidence(obj, action, 0.68, "GeneXus Transaction default maintenance behavior", obj.source_file)
    elif obj.object_type in ("WWP", "WP", "SD", "Query", "Dashboard", "DP"):
        if obj.object_type in ("DP", "Query", "Dashboard"):
            add_evidence(obj, "R", 0.62, f"{obj.object_type} normally reads/provides data", obj.source_file)
        for rgx, actions, reason in WWP_SUFFIX_RULES:
            if rgx.search(obj.object_name):
                for action in actions:
                    add_evidence(obj, action, 0.66, reason, obj.source_file)
                break

    # Name prefixes/suffixes are weaker but helpful.
    lower_name = obj.object_name.lower()
    name_rules = [
        (("get", "find", "query", "search", "list", "load", "view"), "R", 0.58, "name prefix/suffix indicates read"),
        (("create", "insert", "add", "new"), "C", 0.58, "name prefix/suffix indicates create"),
        (("update", "edit", "modify", "set"), "U", 0.58, "name prefix/suffix indicates update"),
        (("delete", "remove"), "D", 0.58, "name prefix/suffix indicates delete"),
    ]
    for words, action, weight, reason in name_rules:
        if lower_name.startswith(words) or lower_name.endswith(words):
            add_evidence(obj, action, weight, reason, obj.source_file, obj.object_name)


def make_feature_names(object_name: str, object_type: str, entities: List[str], actions: Dict[str, float]) -> List[str]:
    entity = entities[0] if entities else strip_object_suffix(object_name)
    active = {a for a, w in actions.items() if w > 0}
    lower = object_name.lower()

    if object_type == "Transaction" or active == {"C", "R", "U", "D"}:
        return [f"{entity} 维护（CRUD）"]
    # When R is only a technical prerequisite for an operation, name the business action.
    if "D" in active and lower.startswith(("delete", "remove")):
        return [f"{entity} 删除"]
    if "U" in active and lower.startswith(("update", "edit", "modify")):
        return [f"{entity} 修改"]
    if "C" in active and lower.startswith(("create", "insert", "add", "new")):
        return [f"{entity} 新增"]
    if active == {"R"}:
        if lower.endswith(("view", "detail", "display")):
            return [f"{entity} 详情查看"]
        return [f"{entity} 列表查询"]
    if active == {"C"}:
        return [f"{entity} 新增"]
    if active == {"U"}:
        return [f"{entity} 修改"]
    if active == {"D"}:
        return [f"{entity} 删除"]
    if active == {"R", "U"}:
        return [f"{entity} 查询修改"]

    names: List[str] = []
    for action in CRUD_ACTIONS:
        if actions.get(action, 0) > 0:
            zh = {"C": "新增", "R": "查询", "U": "修改", "D": "删除"}[action]
            names.append(f"{entity} {zh}")
    return unique_keep_order(names)[:4] or [f"{entity or object_name} 功能"]


def parse_design_sources(design_dir: Path, max_file_bytes: int) -> Tuple[List[DesignObject], List[Dict[str, Any]]]:
    objects: List[DesignObject] = []
    logs: List[Dict[str, Any]] = []
    seen: Dict[Tuple[str, str], DesignObject] = {}

    for src in iter_plain_files(design_dir, max_file_bytes=max_file_bytes):
        if not src.text.strip():
            logs.append({"source": src.path_label, "status": "skip", "message": "empty or unreadable"})
            continue
        extracted = extract_name_type_title(src)
        if not extracted:
            logs.append({"source": src.path_label, "status": "skip", "message": "no GeneXus object name detected"})
            continue
        for name, raw_type, title, desc, hints in extracted:
            guessed = guess_type_from_text(src.text, src.path_label)
            object_type = normalize_object_type(raw_type, guessed, src.path_label, src.text)
            entities = infer_entities(name, object_type, src.text)
            obj = DesignObject(
                object_name=name,
                object_type=object_type,
                source_file=src.path_label,
                title=title,
                description=desc,
                entity=entities[0] if entities else name,
                entities=entities,
                raw_hints=hints,
            )
            infer_actions_from_text(obj, src.text)
            obj.feature_names = make_feature_names(obj.object_name, obj.object_type, obj.entities, obj.actions)

            key = (normalize_key(obj.object_name), obj.object_type)
            if key not in seen:
                seen[key] = obj
            else:
                # Keep richer source; merge evidence/actions/entities.
                cur = seen[key]
                if len(obj.title + obj.description + obj.raw_hints) > len(cur.title + cur.description + cur.raw_hints):
                    cur.title = obj.title or cur.title
                    cur.description = obj.description or cur.description
                    cur.raw_hints = obj.raw_hints or cur.raw_hints
                    cur.source_file = obj.source_file or cur.source_file
                cur.entities = unique_keep_order(cur.entities + obj.entities)
                cur.entity = cur.entities[0] if cur.entities else cur.entity
                for action in CRUD_ACTIONS:
                    cur.actions[action] = max(cur.actions.get(action, 0.0), obj.actions.get(action, 0.0))
                cur.evidence.extend(obj.evidence)
                cur.feature_names = make_feature_names(cur.object_name, cur.object_type, cur.entities, cur.actions)
            logs.append({"source": src.path_label, "status": "parsed", "message": f"{name} ({object_type})"})

    objects = sorted(seen.values(), key=lambda o: (o.object_type, o.object_name.lower()))
    return objects, logs


JAVA_CLASS_RE = re.compile(r"\b(?:public\s+)?(?:final\s+)?(?:class|interface|enum)\s+([A-Za-z_][A-Za-z0-9_]*)\b")
SQL_TABLE_RE = re.compile(r"\b(?:from|join|into|update|delete\s+from)\s+([A-Za-z_][A-Za-z0-9_.$]*)", re.I)


def extract_actions_from_java_text(rel_path: str, text: str) -> Tuple[Dict[str, float], List[Evidence], List[str]]:
    actions = {a: 0.0 for a in CRUD_ACTIONS}
    evs: List[Evidence] = []
    tables: List[str] = []

    for action, patterns in CRUD_PATTERNS.items():
        for rgx, weight, reason in patterns:
            m = rgx.search(text[:1_500_000])
            if m:
                actions[action] = max(actions[action], weight)
                evs.append(Evidence(action, weight, reason, rel_path, compact_snippet(text, m)))

    for m in SQL_TABLE_RE.finditer(text[:1_500_000]):
        table = m.group(1).split(".")[-1]
        if is_probable_name(table):
            tables.append(table)
    return actions, evs, unique_keep_order(tables)


def index_java_files(java_dir: Path, max_file_bytes: int) -> Tuple[List[JavaFileInfo], Dict[str, List[int]], Dict[str, List[int]]]:
    infos: List[JavaFileInfo] = []
    by_stem: Dict[str, List[int]] = defaultdict(list)
    by_token: Dict[str, List[int]] = defaultdict(list)

    for p in sorted(java_dir.rglob("*.java")):
        if any(part in IGNORE_DIR_PARTS for part in p.parts):
            continue
        rel = safe_rel(p, java_dir)
        try:
            text = read_text_safely(p, max_bytes=max_file_bytes)
        except Exception:
            continue
        class_names = JAVA_CLASS_RE.findall(text[:300000])
        tokens = unique_keep_order(camel_tokens(p.stem) + [t for c in class_names for t in camel_tokens(c)])
        actions, evs, tables = extract_actions_from_java_text(rel, text)
        info = JavaFileInfo(
            path=p,
            rel_path=rel,
            stem_norm=normalize_key(p.stem),
            class_names=class_names,
            tokens=tokens,
            actions=actions,
            evidence=evs,
            tables=tables,
            text_sample=text[:20000],
        )
        idx = len(infos)
        infos.append(info)
        by_stem[info.stem_norm].append(idx)
        for tok in set(tokens):
            by_token[tok].append(idx)
    return infos, by_stem, by_token


def match_java_for_object(obj: DesignObject, infos: List[JavaFileInfo], by_stem: Dict[str, List[int]], by_token: Dict[str, List[int]]) -> List[Tuple[float, str, JavaFileInfo]]:
    name = obj.object_name
    exact_candidates = unique_keep_order([name, name.lower(), name.upper(), "a" + name, "A" + name])
    entity_candidate = strip_object_suffix(name)
    exact_norms = [normalize_key(c) for c in exact_candidates if c]
    matches: List[Tuple[float, str, JavaFileInfo]] = []

    for norm in exact_norms:
        for idx in by_stem.get(norm, []):
            matches.append((0.96, "filename_stem_match", infos[idx]))

    if not matches:
        for c in exact_candidates:
            for idx in by_stem.get(normalize_key("a" + c), []):
                matches.append((0.92, "generated_a_prefix_filename_match", infos[idx]))

    if not matches and entity_candidate and normalize_key(entity_candidate) != normalize_key(name):
        # Entity-name file matches are useful traceability hints, but weaker than object-name matches.
        for idx in by_stem.get(normalize_key(entity_candidate), []):
            matches.append((0.45, "entity_filename_match", infos[idx]))

    if not matches:
        token_pool: Set[int] = set()
        for tok in camel_tokens(name)[:2]:
            token_pool.update(by_token.get(tok.lower(), [])[:1000])
        norm_name = normalize_key(name)
        for idx in token_pool:
            info = infos[idx]
            if any(normalize_key(c) == norm_name for c in info.class_names):
                matches.append((0.86, "class_name_match", info))
            elif name in info.text_sample:
                matches.append((0.62, "content_contains_object_name", info))
            elif normalize_key(strip_object_suffix(name)) and normalize_key(strip_object_suffix(name)) in info.stem_norm:
                matches.append((0.45, "stem_contains_entity", info))

    best: Dict[str, Tuple[float, str, JavaFileInfo]] = {}
    for score, method, info in matches:
        if info.rel_path not in best or score > best[info.rel_path][0]:
            best[info.rel_path] = (score, method, info)
    return sorted(best.values(), key=lambda x: (-x[0], x[2].rel_path))[:8]


def enrich_with_java(objects: List[DesignObject], java_dir: Optional[Path], max_file_bytes: int) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    trace_rows: List[Dict[str, Any]] = []
    logs: List[Dict[str, Any]] = []
    if not java_dir:
        return trace_rows, logs
    if not java_dir.exists():
        logs.append({"source": str(java_dir), "status": "warn", "message": "java_dir not found; Java mapping skipped"})
        return trace_rows, logs

    infos, by_stem, by_token = index_java_files(java_dir, max_file_bytes=max_file_bytes)
    logs.append({"source": str(java_dir), "status": "parsed", "message": f"Java files indexed: {len(infos)}"})

    for obj in objects:
        matches = match_java_for_object(obj, infos, by_stem, by_token)
        if not matches:
            trace_rows.append({
                "设计对象类型": obj.object_type,
                "设计对象名": obj.object_name,
                "功能名称(推断)": "; ".join(obj.feature_names),
                "Java文件": "",
                "匹配方式": "not_found",
                "匹配置信度": 0.0,
                "Java CRUD": "",
                "Java表名线索": "",
            })
            continue
        for match_score, method, info in matches:
            obj.java_files.append(info.rel_path)
            obj.java_tables = unique_keep_order(obj.java_tables + info.tables)
            # Java evidence is strong only if mapping is decent.
            for action in CRUD_ACTIONS:
                if info.actions.get(action, 0) > 0:
                    weighted = min(0.98, info.actions[action] * match_score)
                    if match_score >= 0.55 or obj.actions.get(action, 0) > 0:
                        obj.actions[action] = max(obj.actions.get(action, 0), weighted)
            for ev in info.evidence:
                obj.evidence.append(Evidence(ev.action, min(0.98, ev.weight * match_score), f"Java: {ev.reason}", info.rel_path, ev.snippet))
            crud = "".join([a for a in CRUD_ACTIONS if info.actions.get(a, 0) > 0])
            trace_rows.append({
                "设计对象类型": obj.object_type,
                "设计对象名": obj.object_name,
                "功能名称(推断)": "; ".join(obj.feature_names),
                "Java文件": info.rel_path,
                "匹配方式": method,
                "匹配置信度": round(match_score, 2),
                "Java CRUD": crud,
                "Java表名线索": ", ".join(info.tables[:12]),
            })
        obj.java_files = unique_keep_order(obj.java_files)
        obj.java_tables = unique_keep_order(obj.java_tables)
        obj.entities = unique_keep_order(obj.entities + obj.java_tables)
        obj.entity = obj.entities[0] if obj.entities else obj.entity
        obj.feature_names = make_feature_names(obj.object_name, obj.object_type, obj.entities, obj.actions)

    return trace_rows, logs


def has_claude_api() -> bool:
    return bool(os.getenv("ANTHROPIC_API_KEY"))


def has_claude_cli() -> bool:
    return shutil.which("claude") is not None


def parse_json_object_from_text(text: str) -> Optional[Dict[str, Any]]:
    text = text.strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except Exception:
        pass
    m = re.search(r"\{.*\}", text, flags=re.S)
    if m:
        try:
            return json.loads(m.group(0))
        except Exception:
            return None
    return None


def build_claude_prompt(obj: DesignObject) -> str:
    active = "".join(a for a in CRUD_ACTIONS if obj.actions.get(a, 0) > 0) or "none"
    evid = [f"{e.action}:{e.reason}:{e.snippet}" for e in obj.evidence[:8]]
    return textwrap.dedent(f"""
    You are analyzing a GeneXus object to refine a CRUD matrix.
    Return JSON only, no markdown.

    Object name: {obj.object_name}
    Object type: {obj.object_type}
    Current entity candidates: {obj.entities}
    Current CRUD guess: {active}
    Title: {obj.title}
    Description: {obj.description}
    Evidence: {evid}

    Return this schema:
    {{
      "entity": "primary business entity/table",
      "entities": ["entity1"],
      "actions": {{"C": true/false, "R": true/false, "U": true/false, "D": true/false}},
      "feature_names": ["Chinese function name"],
      "confidence": 0.0,
      "reason": "short Chinese reason"
    }}
    """).strip()


def call_claude_cli(prompt: str, timeout: int = 90) -> Optional[Dict[str, Any]]:
    exe = shutil.which("claude")
    if not exe:
        return None
    try:
        proc = subprocess.run([exe, "-p", prompt], capture_output=True, text=True, timeout=timeout)
        if proc.returncode != 0:
            return None
        return parse_json_object_from_text(proc.stdout)
    except Exception:
        return None


def call_claude_api(prompt: str, timeout: int = 90) -> Optional[Dict[str, Any]]:
    if not has_claude_api():
        return None
    try:
        import anthropic  # type: ignore
    except Exception:
        return None
    try:
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        msg = client.messages.create(
            model=os.getenv("ANTHROPIC_MODEL", "claude-3-5-sonnet-latest"),
            max_tokens=800,
            temperature=0,
            messages=[{"role": "user", "content": prompt}],
            timeout=timeout,
        )
        text = "\n".join(block.text for block in msg.content if getattr(block, "type", "") == "text")
        return parse_json_object_from_text(text)
    except Exception:
        return None


def refine_with_claude(objects: List[DesignObject], mode: str, max_objects: int) -> List[Dict[str, Any]]:
    logs: List[Dict[str, Any]] = []
    if mode == "off":
        return logs
    use_api = mode in ("api", "auto") and has_claude_api()
    use_cli = mode in ("cli", "auto") and (has_claude_cli() if not use_api else False)
    if not (use_api or use_cli):
        logs.append({"source": "Claude", "status": "skip", "message": "Claude not available; deterministic parser used"})
        return logs

    ambiguous = [o for o in objects if max(o.actions.values() or [0]) < 0.75 or o.object_type == "Unknown"]
    target = ambiguous[:max_objects]
    for obj in target:
        prompt = build_claude_prompt(obj)
        result = call_claude_api(prompt) if use_api else call_claude_cli(prompt)
        if not result:
            logs.append({"source": obj.object_name, "status": "warn", "message": "Claude returned no usable JSON"})
            continue
        conf = float(result.get("confidence") or 0.75)
        if result.get("entity") and is_probable_name(str(result["entity"])):
            obj.entity = str(result["entity"])
        if isinstance(result.get("entities"), list):
            obj.entities = unique_keep_order([str(x) for x in result["entities"]] + obj.entities)
            obj.entity = obj.entities[0] if obj.entities else obj.entity
        actions = result.get("actions") or {}
        if isinstance(actions, dict):
            for a in CRUD_ACTIONS:
                if bool(actions.get(a)):
                    add_evidence(obj, a, min(0.95, max(0.55, conf)), "Claude refined CRUD inference", "Claude", str(result.get("reason", "")))
        if isinstance(result.get("feature_names"), list) and result["feature_names"]:
            obj.feature_names = unique_keep_order([str(x) for x in result["feature_names"]])
        else:
            obj.feature_names = make_feature_names(obj.object_name, obj.object_type, obj.entities, obj.actions)
        logs.append({"source": obj.object_name, "status": "parsed", "message": f"Claude refined: {''.join(a for a in CRUD_ACTIONS if obj.actions.get(a,0)>0)}"})
    return logs


def evidence_summary(evidence: List[Evidence], limit: int = 5) -> str:
    parts: List[str] = []
    for e in sorted(evidence, key=lambda x: (-x.weight, x.action))[:limit]:
        msg = f"{e.action}:{e.reason}"
        if e.source:
            msg += f"@{e.source}"
        parts.append(msg)
    return " | ".join(parts)


def confidence_label(score: float) -> str:
    if score >= 0.85:
        return "高"
    if score >= 0.65:
        return "中"
    if score > 0:
        return "低"
    return "无"


def build_function_aggregates(objects: List[DesignObject]) -> List[FunctionAggregate]:
    groups: Dict[str, FunctionAggregate] = {}
    for obj in objects:
        features = obj.feature_names or make_feature_names(obj.object_name, obj.object_type, obj.entities, obj.actions)
        for fname in features:
            if fname not in groups:
                groups[fname] = FunctionAggregate(
                    function_id="",
                    function_name=fname,
                    primary_entity=obj.entity,
                    entities=[],
                    actions={a: 0.0 for a in CRUD_ACTIONS},
                    design_objects=[],
                    object_types=[],
                    java_files=[],
                    evidence=[],
                )
            g = groups[fname]
            g.entities = unique_keep_order(g.entities + obj.entities)
            if not g.primary_entity and g.entities:
                g.primary_entity = g.entities[0]
            for a in CRUD_ACTIONS:
                g.actions[a] = max(g.actions.get(a, 0.0), obj.actions.get(a, 0.0))
            g.design_objects = unique_keep_order(g.design_objects + [f"{obj.object_type}:{obj.object_name}"])
            g.object_types = unique_keep_order(g.object_types + [obj.object_type])
            g.java_files = unique_keep_order(g.java_files + obj.java_files)
            g.evidence.extend(obj.evidence)

    # Stable IDs.
    out = sorted(groups.values(), key=lambda g: (g.primary_entity.lower(), g.function_name.lower()))
    for i, g in enumerate(out, start=1):
        g.function_id = f"FUN-{i:04d}"
        if not g.primary_entity and g.entities:
            g.primary_entity = g.entities[0]
    return out


def function_rows(functions: List[FunctionAggregate]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for g in functions:
        max_score = max(g.actions.values() or [0.0])
        rows.append({
            "功能ID": g.function_id,
            "功能名称": g.function_name,
            "主实体/表": g.primary_entity,
            "Create": "Y" if g.actions["C"] > 0 else "",
            "Read": "Y" if g.actions["R"] > 0 else "",
            "Update": "Y" if g.actions["U"] > 0 else "",
            "Delete": "Y" if g.actions["D"] > 0 else "",
            "CRUD组合": "".join(a for a in CRUD_ACTIONS if g.actions.get(a, 0) > 0),
            "置信度": confidence_label(max_score),
            "置信度分数": round(max_score, 2),
            "相关实体/表": ", ".join(g.entities[:20]),
            "设计对象": "; ".join(g.design_objects),
            "对象类型": ", ".join(g.object_types),
            "Java文件数": len(g.java_files),
            "Java文件": "; ".join(g.java_files[:12]),
            "证据摘要": evidence_summary(g.evidence),
        })
    return rows


def object_rows(objects: List[DesignObject]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for obj in objects:
        max_score = max(obj.actions.values() or [0.0])
        rows.append({
            "设计对象类型": obj.object_type,
            "设计对象名": obj.object_name,
            "推断功能名称": "; ".join(obj.feature_names),
            "主实体/表": obj.entity,
            "Create": "Y" if obj.actions["C"] > 0 else "",
            "Read": "Y" if obj.actions["R"] > 0 else "",
            "Update": "Y" if obj.actions["U"] > 0 else "",
            "Delete": "Y" if obj.actions["D"] > 0 else "",
            "CRUD组合": "".join(a for a in CRUD_ACTIONS if obj.actions.get(a, 0) > 0),
            "置信度": confidence_label(max_score),
            "置信度分数": round(max_score, 2),
            "标题/Caption": obj.title,
            "描述/Doc": obj.description,
            "实体候选": ", ".join(obj.entities[:20]),
            "Java文件": "; ".join(obj.java_files[:12]),
            "Java表名线索": ", ".join(obj.java_tables[:20]),
            "来源文件": obj.source_file,
            "证据摘要": evidence_summary(obj.evidence),
        })
    return rows


def entity_rows(functions: List[FunctionAggregate]) -> List[Dict[str, Any]]:
    agg: Dict[str, Dict[str, Any]] = {}
    for g in functions:
        entities = g.entities or ([g.primary_entity] if g.primary_entity else [])
        for ent in entities:
            if ent not in agg:
                agg[ent] = {"实体/表": ent, "C功能数": 0, "R功能数": 0, "U功能数": 0, "D功能数": 0, "功能名称": []}
            for a, col in [("C", "C功能数"), ("R", "R功能数"), ("U", "U功能数"), ("D", "D功能数")]:
                if g.actions.get(a, 0) > 0:
                    agg[ent][col] += 1
            agg[ent]["功能名称"].append(g.function_name)
    rows = []
    for ent, d in agg.items():
        rows.append({
            "实体/表": ent,
            "C功能数": d["C功能数"],
            "R功能数": d["R功能数"],
            "U功能数": d["U功能数"],
            "D功能数": d["D功能数"],
            "CRUD覆盖": "".join(a for a, col in [("C", "C功能数"), ("R", "R功能数"), ("U", "U功能数"), ("D", "D功能数")] if d[col] > 0),
            "相关功能数": len(set(d["功能名称"])),
            "相关功能": "; ".join(sorted(set(d["功能名称"]))[:20]),
        })
    return sorted(rows, key=lambda r: r["实体/表"].lower())


def write_json(out_path: Path, objects: List[DesignObject], functions: List[FunctionAggregate], trace_rows: List[Dict[str, Any]], logs: List[Dict[str, Any]]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "summary": {
            "objects": len(objects),
            "functions": len(functions),
            "trace_rows": len(trace_rows),
        },
        "functions": [asdict(g) for g in functions],
        "objects": [asdict(o) for o in objects],
        "design_java_traceability": trace_rows,
        "parse_log": logs,
    }
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_markdown(out_path: Path, function_table: List[Dict[str, Any]], objects: List[DesignObject], trace_rows: List[Dict[str, Any]], logs: List[Dict[str, Any]]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    counts = Counter(r.get("CRUD组合", "") for r in function_table)
    lines: List[str] = []
    lines.append("# GeneXus 功能 CRUD 表生成结果\n")
    lines.append("## 摘要\n")
    lines.append(f"- 解析设计对象：{len(objects)} 个")
    lines.append(f"- 推断功能：{len(function_table)} 个")
    lines.append(f"- 设计-Java 对照行：{len(trace_rows)} 行")
    lines.append(f"- CRUD组合分布：{', '.join(f'{k or 'None'}={v}' for k, v in sorted(counts.items()))}\n")
    lines.append("## 功能 CRUD 矩阵（前 200 行）\n")
    headers = ["功能ID", "功能名称", "主实体/表", "Create", "Read", "Update", "Delete", "CRUD组合", "置信度", "设计对象"]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for r in function_table[:200]:
        vals = [str(r.get(h, "")).replace("|", "/") for h in headers]
        lines.append("| " + " | ".join(vals) + " |")
    if len(function_table) > 200:
        lines.append(f"\n> 仅显示前 200 行，完整结果请看 Excel/JSON。\n")
    skipped = sum(1 for l in logs if l.get("status") == "skip")
    warned = sum(1 for l in logs if l.get("status") == "warn")
    lines.append("\n## 解析日志摘要\n")
    lines.append(f"- skip: {skipped}")
    lines.append(f"- warn: {warned}")
    out_path.write_text("\n".join(lines), encoding="utf-8")


def write_csv_fallback(out_base: Path, sheets: Dict[str, List[Dict[str, Any]]]) -> None:
    out_base.parent.mkdir(parents=True, exist_ok=True)
    for sheet_name, rows in sheets.items():
        path = out_base.with_name(f"{out_base.stem}_{sheet_name}.csv")
        if not rows:
            path.write_text("", encoding="utf-8")
            continue
        headers = list(rows[0].keys())
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)


def write_excel(out_path: Path, sheets: Dict[str, List[Dict[str, Any]]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        write_csv_fallback(out_path, sheets)
        return False

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        if not rows:
            ws.append(["(empty)"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        # Highlight Y cells for CRUD columns without forcing colors too heavily.
        for idx, h in enumerate(headers, start=1):
            if h in {"Create", "Read", "Update", "Delete"}:
                for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=ws.max_row):
                    for c in cell:
                        c.alignment = Alignment(horizontal="center", vertical="center")
                        if c.value == "Y":
                            c.font = Font(bold=True)
        widths: Dict[int, int] = defaultdict(lambda: 10)
        for row in ws.iter_rows(values_only=True):
            for idx, val in enumerate(row, start=1):
                s = "" if val is None else str(val)
                widths[idx] = max(widths[idx], min(len(s) + 2, 60))
        for idx, width in widths.items():
            header = headers[idx - 1] if idx <= len(headers) else ""
            if header in {"证据摘要", "设计对象", "Java文件", "来源文件", "描述/Doc"}:
                width = min(max(width, 28), 48)
            elif header in {"Create", "Read", "Update", "Delete"}:
                width = 10
            else:
                width = min(max(width, 10), 34)
            ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(out_path)
    return True


def build_outputs(objects: List[DesignObject], functions: List[FunctionAggregate], trace_rows: List[Dict[str, Any]], logs: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    return {
        "Function_CRUD": function_rows(functions),
        "Object_CRUD": object_rows(objects),
        "Entity_Summary": entity_rows(functions),
        "Design_Java": trace_rows,
        "Parse_Log": logs,
    }


def run(args: argparse.Namespace) -> int:
    design_dir = Path(args.design_dir).resolve()
    java_dir = Path(args.java_dir).resolve() if args.java_dir else None
    if not design_dir.exists() or not design_dir.is_dir():
        print(f"[ERROR] design_dir not found or not a directory: {design_dir}", file=sys.stderr)
        return 2

    objects, logs = parse_design_sources(design_dir, max_file_bytes=args.max_file_bytes)
    trace_rows, java_logs = enrich_with_java(objects, java_dir, max_file_bytes=args.max_file_bytes)
    logs.extend(java_logs)
    logs.extend(refine_with_claude(objects, mode=args.claude, max_objects=args.claude_max_objects))

    # Rebuild feature names once after Java/Claude refinement.
    for obj in objects:
        obj.entities = unique_keep_order(obj.entities)
        obj.entity = obj.entities[0] if obj.entities else obj.entity
        obj.feature_names = make_feature_names(obj.object_name, obj.object_type, obj.entities, obj.actions)

    functions = build_function_aggregates(objects)
    sheets = build_outputs(objects, functions, trace_rows, logs)

    out_xlsx = Path(args.out_xlsx).resolve() if args.out_xlsx else None
    out_json = Path(args.out_json).resolve() if args.out_json else None
    out_md = Path(args.out_md).resolve() if args.out_md else None

    if out_xlsx:
        excel_ok = write_excel(out_xlsx, sheets)
        if excel_ok:
            print(f"[OK] Excel written: {out_xlsx}")
        else:
            print(f"[WARN] openpyxl unavailable; CSV fallback written near: {out_xlsx}")
    if out_json:
        write_json(out_json, objects, functions, trace_rows, logs)
        print(f"[OK] JSON written: {out_json}")
    if out_md:
        write_markdown(out_md, sheets["Function_CRUD"], objects, trace_rows, logs)
        print(f"[OK] Markdown written: {out_md}")

    print(f"[OK] Objects={len(objects)} Functions={len(functions)} TraceRows={len(trace_rows)}")
    if not objects:
        print("[WARN] No GeneXus objects detected. Check whether design_dir points to exported XML/XPZ/ZIP/text files.")
    return 0


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="One-shot CRUD table generator for GeneXus systems",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--design-dir", required=True, help="GeneXus export directory containing XML/TXT/JSON/XPZ/ZIP files")
    parser.add_argument("--java-dir", default="", help="Optional generated Java source root, e.g. JavaModel/src/main/java")
    parser.add_argument("--out-xlsx", default="gx_crud_matrix.xlsx", help="Output Excel workbook path")
    parser.add_argument("--out-json", default="gx_crud_matrix.json", help="Output JSON path")
    parser.add_argument("--out-md", default="gx_crud_matrix.md", help="Output Markdown report path")
    parser.add_argument("--max-file-bytes", type=int, default=2_000_000, help="Maximum bytes read per text file")
    parser.add_argument("--claude", choices=["off", "auto", "api", "cli"], default="off", help="Optional Claude refinement mode")
    parser.add_argument("--claude-max-objects", type=int, default=50, help="Maximum ambiguous objects sent to Claude")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    return run(parse_args(argv))


if __name__ == "__main__":
    raise SystemExit(main())
