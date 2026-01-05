#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GeneXus Traceability Builder v3
- Stage 1: Parse design exports (WP/WWP/SD/SDT/Proc/Trn/DP/REST...) -> infer feature names -> Feature_Design/Feature_Group tables
- Stage 2: Scan generated Java sources -> map design object -> java files -> Design_Java table
Output: Excel (.xlsx) with 3 sheets

v3 修正: 支持从单个XML文件中提取多个Object元素
"""

import argparse
import json
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def read_text_safely(path: Path, max_bytes: int = 2_000_000) -> str:
    """Read text with encoding fallbacks; cap size to avoid huge files."""
    data = path.read_bytes()[:max_bytes]
    for enc in ("utf-8", "utf-8-sig", "cp932", "shift_jis", "gbk", "latin-1"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="ignore")


GX_GENERIC_NAMES = {
    # GeneXus の XML 内で頻出する「プロパティ名」や汎用語（これを Object 名として誤検出しがち）
    "name", "value", "description", "desc", "title", "caption",
    "isdefault", "length", "attmaxlen", "attcustomtype", "enumdefinedvalues",
    "properties", "property", "object", "objectname", "objname", "type",
}

def is_probable_name(s: str) -> bool:
    """GeneXus オブジェクト名として妥当かを判定。"""
    if not s:
        return False
    s = s.strip()
    if len(s) > 120:
        return False
    if s.lower() in GX_GENERIC_NAMES:
        return False

    seg_pat = re.compile(r"[A-Za-z_][A-Za-z0-9_]*")
    if "." in s:
        segs = [x for x in s.split(".") if x]
        if not (1 <= len(segs) <= 12):
            return False
        return all(seg_pat.fullmatch(seg) for seg in segs)

    return bool(seg_pat.fullmatch(s))


def normalize_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def camel_tokens(name: str) -> List[str]:
    parts = re.findall(r"[A-Z]+(?=[A-Z][a-z])|[A-Z]?[a-z]+|[0-9]+", name)
    return [p.lower() for p in parts if p]


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


@dataclass
class DesignObject:
    object_name: str
    object_type: str
    source_file: str
    title: str = ""
    description: str = ""
    raw_hints: str = ""
    inferred_entity: str = ""
    inferred_actions: str = ""
    feature_names: str = ""


TYPE_KEYWORDS = {
    "WP": ["web panel", "webpanel", "wp"],
    "WWP": ["workwith", "work with", "workwithplus", "wwp", "ww "],
    "SD": ["smart device", "sd panel", "sdpanel", "mobile panel"],
    "SDT": ["structured data type", "sdt"],
    "Transaction": ["transaction", "trn"],
    "Procedure": ["procedure", "proc"],
    "DP": ["data provider", "dataprovider", "dp"],
    "REST": ["rest", "web service", "webservice", "expose as web service"],
    "API": ["api object", "apiobject"],
    "Theme": ["theme"],
}

ACTION_PATTERNS = [
    (re.compile(r"\b(search|filter|query|find)\b", re.I), "検索"),
    (re.compile(r"\b(list|grid|browse)\b", re.I), "一覧"),
    (re.compile(r"\b(view|detail|show)\b", re.I), "詳細"),
    (re.compile(r"\b(new|create|insert|add)\b", re.I), "新規"),
    (re.compile(r"\b(update|edit|modify)\b", re.I), "更新"),
    (re.compile(r"\b(delete|remove)\b", re.I), "削除"),
    (re.compile(r"\b(export|download)\b", re.I), "エクスポート"),
    (re.compile(r"\b(import|upload)\b", re.I), "インポート"),
    (re.compile(r"\b(approve|approval)\b", re.I), "承認"),
    (re.compile(r"\b(login|signin|auth|authenticate)\b", re.I), "ログイン"),
    (re.compile(r"\b(sync|synchronize)\b", re.I), "同期"),
    (re.compile(r"\b(report|print)\b", re.I), "レポート"),
]

WWP_NAME_RULES = [
    (re.compile(r"(WW|WorkWith)$", re.I), "管理/一覧検索"),
    (re.compile(r"(View|Detail)$", re.I), "詳細表示"),
    (re.compile(r"(Prompt|Select|Picker)$", re.I), "選択/検索"),
    (re.compile(r"(Insert|Create|New)$", re.I), "新規"),
    (re.compile(r"(Update|Edit)$", re.I), "更新"),
    (re.compile(r"(Delete|Remove)$", re.I), "削除"),
]


def guess_type_from_text(text: str, filename: str) -> str:
    t = (text[:20000] + "\n" + filename).lower()
    best = ("Unknown", 0)
    for k, kws in TYPE_KEYWORDS.items():
        score = 0
        for kw in kws:
            if kw in t:
                score += 1
        if score > best[1]:
            best = (k, score)
    return best[0] if best[1] > 0 else "Unknown"


def extract_single_object_from_element(elem: ET.Element, xml_text: str = "") -> Tuple[Optional[str], Optional[str], str, str, str, Dict[str, str]]:
    """从单个XML元素中提取对象信息"""
    props: Dict[str, str] = {}

    def clean(s: str) -> str:
        return re.sub(r"\s+", " ", s or "").strip()

    def add_prop(k: str, v: str) -> None:
        k = clean(k)
        v = clean(v)
        if k and v and k not in props:
            props[k] = v

    attrs = {k: clean(v) for k, v in elem.attrib.items()}
    
    # Properties/Property の Name/Value を辞書化
    for prop in elem.findall(".//Properties/Property"):
        k = prop.findtext("Name") or ""
        v = prop.findtext("Value") or ""
        add_prop(k, v)
    
    # 也检查直接的Property子元素
    for prop in elem.findall("./Property"):
        k = prop.findtext("Name") or prop.findtext("n") or ""
        v = prop.findtext("Value") or prop.findtext("v") or ""
        add_prop(k, v)

    # 候補の組み立て（優先順位つき）
    name_cands: List[str] = []
    name_cands += [attrs.get("fullyQualifiedName", ""), attrs.get("name", ""), attrs.get("Name", "")]
    name_cands += [props.get("Name", ""), props.get("ObjectName", ""), props.get("ObjName", "")]
    name = next((c for c in (clean(x) for x in name_cands) if is_probable_name(c)), None)

    type_cands: List[str] = []
    type_cands += [attrs.get("objectType", ""), attrs.get("typeName", ""), attrs.get("ObjectType", ""), attrs.get("Type", "")]
    type_cands += [props.get("ObjectType", ""), props.get("Type", "")]
    otype = next((c for c in (clean(x) for x in type_cands) if c), None)

    title = clean(props.get("Title", "") or props.get("Caption", "") or attrs.get("title", ""))
    desc = clean(attrs.get("description", "") or props.get("Description", "") or props.get("Desc", ""))

    # 从元素文本内容中提取关键词
    elem_text = ET.tostring(elem, encoding='unicode', method='text') if elem is not None else ""
    combined_text = elem_text + " " + xml_text[:5000]
    
    keywords = re.findall(r"\b(Search|Filter|Grid|For\s+each|Link|Call|Export|Import|Approve|Login|Sync|REST|SDT|DataProvider)\b",
                          combined_text, flags=re.I)
    hints = ",".join(sorted(set(k.lower() for k in keywords)))[:300] if keywords else ""
    if "EnumDefinedValues" in props:
        hints = (hints + ("," if hints else "") + "enum")[:300]

    return name, otype, title, desc, hints, props


def extract_all_objects_from_xml(xml_text: str) -> List[Tuple[Optional[str], Optional[str], str, str, str, Dict[str, str]]]:
    """从XML中提取所有Object元素的信息 - 这是关键修复"""
    xml_text = xml_text.strip()
    results = []

    root = None
    try:
        root = ET.fromstring(xml_text)
    except Exception:
        # 如果解析失败，尝试用正则表达式提取
        return extract_objects_by_regex(xml_text)

    if root is None:
        return []

    # 查找所有可能包含对象定义的元素
    # GeneXus XML 可能有多种结构
    object_elements = []
    
    # 1. 查找 <Object> 元素（最常见）
    object_elements.extend(root.findall(".//Object"))
    
    # 2. 查找 <Objects/Object> 结构
    for objects_container in root.findall(".//Objects"):
        object_elements.extend(objects_container.findall("Object"))
    
    # 3. 如果根元素本身就是Object
    if root.tag == "Object":
        object_elements.append(root)
    
    # 4. 查找 KnowledgeBase/Objects/Object 结构
    for kb in root.findall(".//KnowledgeBase"):
        for objects_container in kb.findall("Objects"):
            object_elements.extend(objects_container.findall("Object"))
    
    # 5. 查找其他可能的对象容器 (Export/Objects, etc.)
    for container in root.findall(".//*"):
        if container.tag in ("Objects", "Export", "Model", "KB"):
            object_elements.extend(container.findall(".//Object"))

    # 去重（基于元素的id或位置）
    seen_elements = set()
    unique_elements = []
    for elem in object_elements:
        elem_id = id(elem)
        if elem_id not in seen_elements:
            seen_elements.add(elem_id)
            unique_elements.append(elem)

    # 如果没有找到Object元素，尝试将根元素作为单个对象处理
    if not unique_elements:
        result = extract_single_object_from_element(root, xml_text)
        if result[0]:  # 如果找到了名称
            return [result]
        return []

    # 处理每个Object元素
    for elem in unique_elements:
        result = extract_single_object_from_element(elem, "")
        if result[0]:  # 如果找到了有效的名称
            results.append(result)

    return results


def extract_objects_by_regex(xml_text: str) -> List[Tuple[Optional[str], Optional[str], str, str, str, Dict[str, str]]]:
    """当XML解析失败时，使用正则表达式提取对象"""
    results = []
    
    # 匹配 <Object ...> 标签
    object_pattern = re.compile(
        r'<\s*Object\s+([^>]*?)(?:/>|>.*?</\s*Object\s*>)',
        re.IGNORECASE | re.DOTALL
    )
    
    for match in object_pattern.finditer(xml_text):
        attr_text = match.group(1)
        full_text = match.group(0)
        
        props: Dict[str, str] = {}
        
        # 提取属性
        attrs = {}
        for k, v in re.findall(r'\b([A-Za-z_][A-Za-z0-9_]*)\s*=\s*"([^"]*)"', attr_text):
            attrs[k] = v
        
        # 提取 Properties
        for k, v in re.findall(
            r'<\s*Property\s*>\s*<\s*Name\s*>\s*([^<]+?)\s*</\s*Name\s*>\s*<\s*Value\s*>\s*(.*?)\s*</\s*Value\s*>\s*</\s*Property\s*>',
            full_text, flags=re.IGNORECASE | re.DOTALL
        ):
            props[k.strip()] = re.sub(r'\s+', ' ', v).strip()

        def clean(s: str) -> str:
            return re.sub(r"\s+", " ", s or "").strip()

        name_cands = [attrs.get("fullyQualifiedName", ""), attrs.get("name", ""), 
                      props.get("Name", ""), props.get("ObjectName", "")]
        name = next((c for c in (clean(x) for x in name_cands) if is_probable_name(c)), None)

        type_cands = [attrs.get("objectType", ""), attrs.get("typeName", ""),
                      props.get("ObjectType", ""), props.get("Type", "")]
        otype = next((c for c in (clean(x) for x in type_cands) if c), None)

        title = clean(props.get("Title", "") or props.get("Caption", ""))
        desc = clean(attrs.get("description", "") or props.get("Description", ""))

        if name:
            results.append((name, otype, title, desc, "", props))

    return results


def extract_name_type_title_from_xml(xml_text: str) -> Tuple[Optional[str], Optional[str], str, str, str, Dict[str, str]]:
    """保持向后兼容的单对象提取函数"""
    results = extract_all_objects_from_xml(xml_text)
    if results:
        return results[0]
    return None, None, "", "", "", {}


def parse_enum_defined_values(raw: str) -> List[str]:
    """EnumDefinedValues の Value から「表示ラベル」部分を抽出する。"""
    raw = re.sub(r"\s+", " ", raw or "").strip()
    if not raw:
        return []
    parts = re.split(r"[：:]\s*(?=\d+\.)", raw)
    out: List[str] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        m = re.match(r"\d+\.\s*[^,]{1,200},\s*(.+)$", p)
        label = m.group(1).strip() if m else p
        label = label.strip(" ：:;")
        if label and label not in out:
            out.append(label)
    return out


def extract_name_type_title_from_text(text: str) -> Tuple[Optional[str], Optional[str], str, str, str]:
    name = None
    m = re.search(r"^\s*(?:Object|Name)\s*[:=]\s*([A-Za-z_][A-Za-z0-9_]*)\s*$", text, flags=re.M)
    if m:
        name = m.group(1)
    else:
        m = re.search(r"^\s*(Web\s*Panel|Transaction|Procedure|Data\s*Provider|SDT|Smart\s*Device)\s*[:=]\s*([A-Za-z_][A-Za-z0-9_]*)",
                      text, flags=re.I | re.M)
        if m:
            name = m.group(2)

    otype = None
    m = re.search(r"^\s*Type\s*[:=]\s*(.+?)\s*$", text, flags=re.I | re.M)
    if m:
        otype = m.group(1).strip()

    title = ""
    m = re.search(r"^\s*(?:Title|Caption)\s*[:=]\s*(.+?)\s*$", text, flags=re.I | re.M)
    if m:
        title = m.group(1).strip()

    desc = ""
    m = re.search(r"^\s*(?:Description|Doc|Documentation)\s*[:=]\s*(.+?)\s*$", text, flags=re.I | re.M)
    if m:
        desc = m.group(1).strip()

    keywords = re.findall(r"\b(Search|Filter|Grid|For\s+each|Link|Call|Export|Import|Approve|Login|Sync|REST)\b",
                          text, flags=re.I)
    hints = ",".join(sorted(set(k.lower() for k in keywords)))[:300] if keywords else ""
    return name, otype, title, desc, hints


def normalize_object_type(raw_type: Optional[str], guessed: str, filename: str, text: str) -> str:
    rt = (raw_type or "").strip().lower()
    if "web panel" in rt or rt in ("wp", "webpanel"):
        return "WP"
    if "workwith" in rt or "work with" in rt or "workwithplus" in rt or rt == "wwp":
        return "WWP"
    if "smart device" in rt or "sd panel" in rt or "sdpanel" in rt:
        return "SD"
    if "structured data type" in rt or rt == "sdt":
        return "SDT"
    if "transaction" in rt or rt == "trn":
        return "Transaction"
    if "procedure" in rt or rt == "proc":
        return "Procedure"
    if "data provider" in rt or rt == "dp" or "dataprovider" in rt:
        return "DP"
    if "api object" in rt or "apiobject" in rt:
        return "API"
    if "theme" in rt:
        return "Theme"

    if guessed in ("WP", "WWP", "SD", "SDT", "Transaction", "Procedure", "DP", "REST", "API", "Theme"):
        return guessed

    fn = filename.lower()
    if "workwith" in fn or fn.endswith("ww.xml") or fn.endswith("wwp.xml"):
        return "WWP"
    if "sdt" in fn:
        return "SDT"
    if "trn" in fn or "transaction" in fn:
        return "Transaction"
    if "proc" in fn or "procedure" in fn:
        return "Procedure"
    if "dp" in fn or "dataprovider" in fn:
        return "DP"
    if "sd" in fn and ("panel" in fn or "smart" in text.lower()):
        return "SD"
    return "Unknown"


def infer_entity(object_name: str, text: str) -> str:
    base_name = (object_name.split(".")[-1] if object_name else object_name)

    m = re.search(r"\b(Transaction|Trn)\s*[:=]\s*([A-Za-z_][A-Za-z0-9_]*)", text, flags=re.I)
    if m:
        return m.group(2)

    name = base_name
    for suf in ("WW", "View", "Detail", "Prompt", "Picker", "List", "Grid", "SD", "Panel"):
        if name.lower().endswith(suf.lower()) and len(name) > len(suf) + 1:
            name = name[: -len(suf)]
            break
    m = re.match(r"WorkWith(.+)$", name, flags=re.I)
    if m and is_probable_name(m.group(1)):
        return m.group(1)
    return name


def infer_actions(object_name: str, object_type: str, text: str, title: str) -> List[str]:
    actions: List[str] = []

    if object_type == "WWP":
        for rgx, act in WWP_NAME_RULES:
            if rgx.search(object_name):
                actions.extend(act.split("/"))
                break

    if object_type in ("WP", "SD"):
        if re.search(r"\b(grid|for\s+each)\b", text, flags=re.I):
            actions.append("一覧")
        if re.search(r"\b(search|filter|where\s+.*like)\b", text, flags=re.I):
            actions.append("検索")

    if object_type == "Transaction":
        actions.extend(["新規", "更新", "削除"])

    if object_type in ("Procedure", "DP", "REST", "API"):
        name_l = object_name.lower()
        if name_l.startswith(("get", "find", "query")):
            actions.append("検索")
        if name_l.startswith(("list", "search", "load")):
            actions.extend(["一覧", "検索"])
        if name_l.startswith(("create", "insert", "add", "new")):
            actions.append("新規")
        if name_l.startswith(("update", "edit", "set")):
            actions.append("更新")
        if name_l.startswith(("delete", "remove")):
            actions.append("削除")
        if name_l.startswith(("export", "download")):
            actions.append("エクスポート")
        if name_l.startswith(("import", "upload")):
            actions.append("インポート")
        if name_l.startswith(("sync", "push", "pull")):
            actions.append("同期")
        if name_l.startswith(("login", "auth")):
            actions.append("ログイン")
        combined_local = f"{title}\n{text}"
        if re.search(r"(帳票|発行|印刷|出力|レポート|PDF|Excel|CSV)", combined_local):
            actions.append("出力")
        if re.search(r"(連携|同期|インタフェース|インターフェース|I/F|IF|送信|受信|取込|取り込み)", combined_local):
            actions.append("連携")

    combined = f"{title}\n{text}"
    for rgx, act in ACTION_PATTERNS:
        if rgx.search(combined):
            actions.append(act)

    out, seen = [], set()
    for a in actions:
        if a and a not in seen:
            out.append(a)
            seen.add(a)

    if not out:
        out = ["データモデル"] if object_type == "SDT" else ["機能"]
    return out


def make_feature_names(entity: str, object_type: str, actions: List[str]) -> List[str]:
    if "一覧" in actions and "検索" in actions:
        return [f"{entity} 一覧検索"]
    if object_type == "Transaction" and all(x in actions for x in ("新規", "更新", "削除")):
        return [f"{entity} メンテナンス（新規/更新/削除）"]

    def has_ja(s: str) -> bool:
        return any(ord(ch) > 127 for ch in s)

    uniq_actions: List[str] = []
    for a in actions:
        if a not in uniq_actions:
            uniq_actions.append(a)

    if entity and has_ja(entity):
        if re.search(r"(発行|帳票|出力|印刷|レポート)", entity):
            return [entity]
        if uniq_actions == ["機能"]:
            return [entity]

    res: List[str] = []
    for act in uniq_actions:
        if act == "詳細":
            res.append(f"{entity} 詳細表示")
        elif act == "データモデル":
            res.append(f"{entity} データモデル")
        elif act == "機能":
            res.append(f"{entity} 機能")
        else:
            res.append(f"{entity} {act}")

    out, seen = [], set()
    for x in res:
        if x not in seen:
            out.append(x)
            seen.add(x)
    return out[:3]


def create_design_object_from_extracted(
    name: str, 
    otype: Optional[str], 
    title: str, 
    desc: str, 
    hints: str, 
    props: Dict[str, str],
    source_label: str,
    filename: str,
    text: str
) -> Optional[DesignObject]:
    """从提取的数据创建DesignObject"""
    if not name or not is_probable_name(name):
        return None
    
    guessed = guess_type_from_text(text, filename)
    object_type = normalize_object_type(otype, guessed, filename, text)
    
    entity = infer_entity(name, text)
    actions = infer_actions(name, object_type, text, title)

    enum_labels = parse_enum_defined_values(props.get("EnumDefinedValues", ""))
    display_entity = (desc or title or entity or name)

    if enum_labels:
        features = [f"{display_entity}：{lab}" for lab in enum_labels][:30]
    else:
        features = make_feature_names(display_entity, object_type, actions)

    return DesignObject(
        object_name=name, 
        object_type=object_type, 
        source_file=source_label,
        title=title, 
        description=desc, 
        raw_hints=hints,
        inferred_entity=entity, 
        inferred_actions=",".join(actions),
        feature_names=",".join(features),
    )


def parse_design_file(path: Path, source_label: str) -> List[DesignObject]:
    """解析设计文件 - 修复版：支持提取多个对象"""
    text = read_text_safely(path)
    filename = path.name.lower()
    results: List[DesignObject] = []

    # XML 文件处理
    if filename.endswith((".xml", ".gxobj", ".gxd", ".gxl")) or text.lstrip().startswith("<"):
        # 使用新的多对象提取函数
        extracted_objects = extract_all_objects_from_xml(text)
        
        if extracted_objects:
            for name, otype, title, desc, hints, props in extracted_objects:
                obj = create_design_object_from_extracted(
                    name, otype, title, desc, hints, props,
                    source_label, path.name, text
                )
                if obj:
                    results.append(obj)
        
        # 如果没有提取到任何对象，尝试从文件名获取
        if not results:
            stem = path.stem
            if is_probable_name(stem):
                guessed = guess_type_from_text(text, path.name)
                object_type = normalize_object_type(None, guessed, path.name, text)
                entity = infer_entity(stem, text)
                actions = infer_actions(stem, object_type, text, "")
                features = make_feature_names(entity, object_type, actions)
                results.append(DesignObject(
                    object_name=stem, 
                    object_type=object_type, 
                    source_file=source_label,
                    title="", 
                    description="", 
                    raw_hints="filename_fallback",
                    inferred_entity=entity, 
                    inferred_actions=",".join(actions),
                    feature_names=",".join(features),
                ))
        
        return results

    # JSON 文件处理
    if filename.endswith(".json"):
        try:
            data = json.loads(text)
            # 支持JSON数组
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict):
                        name = item.get("ObjectName") or item.get("Name")
                        otype = item.get("ObjectType") or item.get("Type")
                        title = item.get("Title") or item.get("Caption") or ""
                        desc = item.get("Description") or ""
                        if name and is_probable_name(name):
                            guessed = guess_type_from_text(str(item), path.name)
                            object_type = normalize_object_type(otype, guessed, path.name, str(item))
                            entity = infer_entity(name, str(item))
                            actions = infer_actions(name, object_type, str(item), title)
                            features = make_feature_names(entity, object_type, actions)
                            results.append(DesignObject(
                                object_name=name, 
                                object_type=object_type, 
                                source_file=source_label,
                                title=title, 
                                description=desc, 
                                raw_hints="json",
                                inferred_entity=entity, 
                                inferred_actions=",".join(actions),
                                feature_names=",".join(features),
                            ))
            elif isinstance(data, dict):
                # 检查是否有Objects数组
                objects_list = data.get("Objects") or data.get("objects") or [data]
                if not isinstance(objects_list, list):
                    objects_list = [data]
                for item in objects_list:
                    if isinstance(item, dict):
                        name = item.get("ObjectName") or item.get("Name")
                        otype = item.get("ObjectType") or item.get("Type")
                        title = item.get("Title") or item.get("Caption") or ""
                        desc = item.get("Description") or ""
                        if name and is_probable_name(name):
                            guessed = guess_type_from_text(str(item), path.name)
                            object_type = normalize_object_type(otype, guessed, path.name, str(item))
                            entity = infer_entity(name, str(item))
                            actions = infer_actions(name, object_type, str(item), title)
                            features = make_feature_names(entity, object_type, actions)
                            results.append(DesignObject(
                                object_name=name, 
                                object_type=object_type, 
                                source_file=source_label,
                                title=title, 
                                description=desc, 
                                raw_hints="json",
                                inferred_entity=entity, 
                                inferred_actions=",".join(actions),
                                feature_names=",".join(features),
                            ))
            if results:
                return results
        except Exception:
            pass

    # 纯文本文件处理
    name, otype, title, desc, hints = extract_name_type_title_from_text(text)
    guessed = guess_type_from_text(text, path.name)
    object_type = normalize_object_type(otype, guessed, path.name, text)

    if not (name and is_probable_name(name)):
        stem = path.stem
        name = stem if is_probable_name(stem) else None

    if name:
        entity = infer_entity(name, text)
        actions = infer_actions(name, object_type, text, title)
        features = make_feature_names(entity, object_type, actions)
        results.append(DesignObject(
            object_name=name, 
            object_type=object_type, 
            source_file=source_label,
            title=title, 
            description=desc, 
            raw_hints=hints or "filename",
            inferred_entity=entity, 
            inferred_actions=",".join(actions),
            feature_names=",".join(features),
        ))
    
    return results


def iter_design_files(design_dir: Path) -> List[Tuple[Path, str]]:
    items: List[Tuple[Path, str]] = []
    extract_root = design_dir / ".tmp_extract"
    if extract_root.exists():
        for old in extract_root.rglob("*"):
            if old.is_file():
                try:
                    old.unlink()
                except Exception:
                    pass

    for p in design_dir.rglob("*"):
        if p.is_dir():
            continue
        lower = p.name.lower()
        if lower.endswith((".xpz", ".zip")):
            out_dir = extract_root / (p.stem + "_extracted")
            ensure_dir(out_dir)
            try:
                with zipfile.ZipFile(p, "r") as zf:
                    zf.extractall(out_dir)
                for q in out_dir.rglob("*"):
                    if q.is_file():
                        items.append((q, f"archive:{p.name}:{q.relative_to(out_dir)}"))
            except Exception:
                continue
        else:
            items.append((p, str(p.relative_to(design_dir))))
    return items


def build_design_objects(design_dir: Path) -> List[DesignObject]:
    objs: List[DesignObject] = []
    for p, label in iter_design_files(design_dir):
        if p.suffix.lower() in (".png", ".jpg", ".jpeg", ".gif", ".pdf", ".docx", ".xlsx", ".pptx"):
            continue
        try:
            objs.extend(parse_design_file(p, label))
        except Exception as e:
            print(f"[WARN] Failed to parse {label}: {e}", file=sys.stderr)
            continue

    # 去重：保留信息最丰富的版本
    best: Dict[Tuple[str, str], DesignObject] = {}
    for o in objs:
        key = (o.object_name, o.object_type)
        if key not in best:
            best[key] = o
        else:
            cur = best[key]
            score_cur = len(cur.title) + len(cur.description) + len(cur.raw_hints)
            score_new = len(o.title) + len(o.description) + len(o.raw_hints)
            if score_new > score_cur:
                best[key] = o
    return list(best.values())


def build_feature_design_rows(objs: List[DesignObject]) -> List[Dict]:
    rows: List[Dict] = []
    for o in objs:
        feats = [x.strip() for x in o.feature_names.split(",") if x.strip()] or [f"{o.inferred_entity or o.object_name} 機能"]
        for f in feats:
            rows.append({
                "機能名": f,
                "設計オブジェクト種別": o.object_type,
                "設計オブジェクト名": o.object_name,
                "エンティティ(推定)": o.inferred_entity,
                "アクション(推定)": o.inferred_actions,
                "タイトル/Caption": o.title,
                "説明/Doc": o.description,
                "ヒント": o.raw_hints,
                "元ファイル": o.source_file,
            })
    return rows


def build_feature_group_rows(feature_design_rows: List[Dict]) -> List[Dict]:
    group: Dict[str, List[Tuple[str, str]]] = {}
    for r in feature_design_rows:
        group.setdefault(r["機能名"], []).append((r["設計オブジェクト種別"], r["設計オブジェクト名"]))

    out: List[Dict] = []
    for f, lst in sorted(group.items(), key=lambda x: x[0]):
        objs = "; ".join([f"{t}:{n}" for t, n in sorted(set(lst))])
        out.append({"機能名": f, "対応設計オブジェクト": objs, "オブジェクト数": len(set(lst))})
    return out


JAVA_CLASS_RE = re.compile(r"^\s*(public\s+)?(final\s+)?class\s+([A-Za-z_][A-Za-z0-9_]*)\b", re.M)


def index_java_files(java_dir: Path) -> Tuple[Dict[str, List[Path]], Dict[str, List[Path]]]:
    by_stem: Dict[str, List[Path]] = {}
    by_token: Dict[str, List[Path]] = {}
    for p in java_dir.rglob("*.java"):
        sn = normalize_key(p.stem)
        by_stem.setdefault(sn, []).append(p)
        for tok in set(camel_tokens(p.stem)):
            by_token.setdefault(tok, []).append(p)
    return by_stem, by_token


def match_java_for_object(obj: DesignObject,
                          by_stem: Dict[str, List[Path]],
                          by_token: Dict[str, List[Path]],
                          java_dir: Path) -> List[Dict]:
    name = obj.object_name
    candidates = [name, name.lower(), name.upper(), "a" + name, "A" + name]
    cand_norms = [normalize_key(c) for c in candidates if c]

    matches: List[Tuple[float, str, Path]] = []

    for cn in cand_norms:
        for p in by_stem.get(cn, []):
            matches.append((0.95, "filename_stem_match", p))

    if not matches:
        toks = camel_tokens(name)
        primary = toks[0] if toks else normalize_key(name)
        pool = by_token.get(primary, [])[:2000]
        for p in pool:
            try:
                head = read_text_safely(p, max_bytes=200_000)
            except Exception:
                continue
            cm = JAVA_CLASS_RE.search(head)
            if cm and normalize_key(cm.group(3)) == normalize_key(name):
                matches.append((0.85, "class_name_match_in_file", p))
                continue
            if name in head:
                matches.append((0.60, "content_contains_object_name", p))

    if not matches:
        nn = normalize_key(name)
        for stem_norm, paths in by_stem.items():
            if nn and nn in stem_norm:
                for p in paths[:5]:
                    matches.append((0.40, "stem_contains_name", p))
            if len(matches) >= 10:
                break

    best: Dict[str, Tuple[float, str, Path]] = {}
    for sc, method, p in matches:
        k = str(p)
        if k not in best or sc > best[k][0]:
            best[k] = (sc, method, p)

    ranked = sorted(best.values(), key=lambda x: (-x[0], x[2].name))[:5]
    out = [{
        "設計オブジェクト種別": obj.object_type,
        "設計オブジェクト名": obj.object_name,
        "機能名(推定)": obj.feature_names,
        "Javaファイル": str(p.relative_to(java_dir)),
        "マッチ方法": method,
        "信頼度": round(sc, 2),
    } for sc, method, p in ranked]

    if not out:
        out.append({
            "設計オブジェクト種別": obj.object_type,
            "設計オブジェクト名": obj.object_name,
            "機能名(推定)": obj.feature_names,
            "Javaファイル": "",
            "マッチ方法": "not_found",
            "信頼度": 0.0,
        })
    return out


def build_design_java_rows(objs: List[DesignObject], java_src_dir: Path) -> List[Dict]:
    by_stem, by_token = index_java_files(java_src_dir)
    rows: List[Dict] = []
    for o in objs:
        if o.object_type == "Theme":
            continue
        rows.extend(match_java_for_object(o, by_stem, by_token, java_src_dir))
    return rows


def autosize_worksheet(ws, max_width: int = 60) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for i, val in enumerate(row, start=1):
            s = "" if val is None else str(val)
            widths[i] = max(widths.get(i, 0), len(s))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(10, w + 2), max_width)


def write_sheet(wb: Workbook, title: str, rows: List[Dict]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["(empty)"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    autosize_worksheet(ws)


def save_excel(out_path: Path,
               feature_design_rows: List[Dict],
               feature_group_rows: List[Dict],
               design_java_rows: List[Dict]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Feature_Design", feature_design_rows)
    write_sheet(wb, "Feature_Group", feature_group_rows)
    write_sheet(wb, "Design_Java", design_java_rows)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="GeneXus: Feature-Design & Design-Java traceability tables (v3)")
    ap.add_argument("--design_dir", required=True, help="Folder containing exported design files (xml/txt/xpz/zip)")
    ap.add_argument("--java_dir", required=False, default="", help="Java source root, e.g. JavaModel/src/main/java")
    ap.add_argument("--out", required=True, help="Output Excel file path (.xlsx)")
    ap.add_argument("--dump_objects_json", default="", help="Optional: dump parsed objects to JSON")
    ap.add_argument("--verbose", "-v", action="store_true", help="Show verbose output")
    args = ap.parse_args()

    design_dir = Path(args.design_dir).resolve()
    if not design_dir.exists():
        print(f"[ERROR] design_dir not found: {design_dir}", file=sys.stderr)
        sys.exit(2)

    out_path = Path(args.out).resolve()
    ensure_dir(out_path.parent)

    print(f"[INFO] Scanning design files in: {design_dir}")
    objs = build_design_objects(design_dir)
    objs = sorted(objs, key=lambda o: (o.object_type, o.object_name))
    print(f"[INFO] Found {len(objs)} unique design objects")

    if args.verbose:
        type_counts: Dict[str, int] = {}
        for o in objs:
            type_counts[o.object_type] = type_counts.get(o.object_type, 0) + 1
        for t, c in sorted(type_counts.items()):
            print(f"  - {t}: {c}")

    feature_design_rows = build_feature_design_rows(objs)
    feature_group_rows = build_feature_group_rows(feature_design_rows)

    if args.dump_objects_json:
        dump_path = Path(args.dump_objects_json).resolve()
        ensure_dir(dump_path.parent)
        dump_path.write_text(json.dumps([asdict(o) for o in objs], ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[INFO] Dumped objects to: {dump_path}")

    design_java_rows: List[Dict] = []
    if args.java_dir:
        java_dir = Path(args.java_dir).resolve()
        if not java_dir.exists():
            print(f"[WARN] java_dir not found, skip java mapping: {java_dir}", file=sys.stderr)
        else:
            design_java_rows = build_design_java_rows(objs, java_dir)

    save_excel(out_path, feature_design_rows, feature_group_rows, design_java_rows)
    print(f"[OK] Written: {out_path}")
    print(f"[OK] Objects: {len(objs)} | Feature rows: {len(feature_design_rows)} | Java rows: {len(design_java_rows)}")


if __name__ == "__main__":
    main()
