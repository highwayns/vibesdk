#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel → Markdown converter (single-file script)

Key capabilities:
- Supports .xlsx (openpyxl) and .xls (xlrd)
- Expands merged cells
- Splits into table/notes blocks and outputs Markdown
- Attempts structured extraction for "ロジック設計書テンプレート" sheets
- Extracts embedded raster images from .xlsx (openpyxl)
- Attempts DrawingML (flowchart/shapes) → Mermaid for .xlsx
- Generates: Title / Overview / Document Info / Update History / per-sheet sections
"""

from __future__ import annotations

import argparse
import datetime as _dt
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

# Optional deps
try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None

try:
    import xlrd  # type: ignore
except Exception:
    xlrd = None

try:
    from PIL import Image  # type: ignore
    from io import BytesIO
except Exception:
    Image = None
    BytesIO = None


def eprint(*args: Any) -> None:
    print(*args, file=sys.stderr)


def is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def norm_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return str(v).strip()


def safe_md(s: str) -> str:
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("|", "\\|")
    return s


def slugify(s: str, max_len: int = 64) -> str:
    s0 = re.sub(r"\s+", "_", s.strip())
    s0 = re.sub(r"[^\w\-\u4e00-\u9fff\u3040-\u30ff\u3400-\u4dbf]", "_", s0)
    s0 = re.sub(r"_+", "_", s0).strip("_")
    return (s0[:max_len] or "sheet")


def detect_management_no(filename: str) -> Optional[str]:
    m = re.search(r"\b[A-Z]{2}-\d{6}-\d{3}(?:\.\d+)?\b", filename)
    if m:
        return m.group(0)
    m = re.search(r"\b[A-Z]{2}-\d{3,}(?:-\d{2,})*(?:\.\d+)?\b", filename)
    return m.group(0) if m else None


def guess_doc_classification(title: str, sheet_names: Sequence[str]) -> str:
    t = title
    s = " ".join(sheet_names)
    for kw, cls in [
        ("Dao", "Dao設計書"),
        ("DAO", "Dao設計書"),
        ("業務ロジック", "業務ロジック設計書"),
        ("ロジック", "業務ロジック設計書"),
        ("画面", "画面設計書"),
        ("帳票", "帳票設計書"),
    ]:
        if kw in t or kw in s:
            return cls
    return "-"


def clean_header_row(row: List[str]) -> List[str]:
    cleaned: List[str] = []
    prev = None
    for v in row:
        if v == prev and v != "":
            cleaned.append("")
        else:
            cleaned.append(v)
            prev = v
    return cleaned


def find_actual_header_row(data: List[List[str]], sheet_name: str) -> int:
    for i, row in enumerate(data):
        uniq = set(v for v in row if v)
        if len(uniq) > 1 or (len(uniq) == 1 and sheet_name not in uniq):
            return i
    return 0


def trim_grid(grid: List[List[str]]) -> List[List[str]]:
    if not grid:
        return grid
    rows = [r for r in grid if any(not is_blank(c) for c in r)]
    if not rows:
        return []
    maxc = max(len(r) for r in rows)
    rows = [r + [""] * (maxc - len(r)) for r in rows]
    non_empty_cols = [any(not is_blank(rows[r][c]) for r in range(len(rows))) for c in range(maxc)]
    if any(non_empty_cols):
        kept_idx = [i for i, ok in enumerate(non_empty_cols) if ok]
        rows = [[row[i] for i in kept_idx] for row in rows]
    while rows and rows[0]:
        last = len(rows[0]) - 1
        if all(is_blank(r[last]) for r in rows):
            rows = [r[:-1] for r in rows]
        else:
            break
    return rows


def row_nonempty_count(row: List[str]) -> int:
    return sum(1 for c in row if not is_blank(c))


def split_blocks_by_empty_rows(grid: List[List[str]], empty_run: int = 2) -> List[List[List[str]]]:
    blocks: List[List[List[str]]] = []
    cur: List[List[str]] = []
    blanks = 0
    for row in grid:
        if row_nonempty_count(row) == 0:
            blanks += 1
            if cur and blanks >= empty_run:
                blocks.append(cur)
                cur = []
        else:
            blanks = 0
            cur.append(row)
    if cur:
        blocks.append(cur)
    return blocks


def is_table_block(block: List[List[str]]) -> bool:
    if len(block) < 2:
        return False
    counts = [row_nonempty_count(r) for r in block]
    avg = sum(counts) / max(1, len(counts))
    return avg >= 2.0


def md_table(rows: List[List[str]]) -> str:
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    header = clean_header_row([safe_md(norm_text(v)) for v in rows[0]])
    body = [[safe_md(norm_text(v)) for v in r] for r in rows[1:]]
    if all(h == "" for h in header):
        header = [f"col{i+1}" for i in range(width)]
    lines = []
    lines.append("| " + " | ".join(header) + " |")
    lines.append("|" + "|".join(["---"] * width) + "|")
    for r in body:
        lines.append("| " + " | ".join(r + [""] * (width - len(r))) + " |")
    return "\n".join(lines)


def md_paragraph_block(block: List[List[str]]) -> str:
    lines: List[str] = []
    for r in block:
        parts = [norm_text(c) for c in r if not is_blank(c)]
        if parts:
            lines.append(" ".join(parts))
    return "\n\n".join(lines)


# -----------------------------
# .xlsx reader
# -----------------------------
def load_xlsx_grid(path: Path, data_only: bool = True) -> Tuple[List[str], Dict[str, List[List[str]]], List[str]]:
    diffs: List[str] = []
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Please install: pip install openpyxl")
    wb = load_workbook(str(path), data_only=data_only)
    sheet_names = wb.sheetnames
    sheets: Dict[str, List[List[str]]] = {}
    for ws in wb.worksheets:
        try:
            merged_ranges = list(ws.merged_cells.ranges)
            for rng in merged_ranges:
                tl = ws.cell(rng.min_row, rng.min_col).value
                ws.unmerge_cells(str(rng))
                for r in range(rng.min_row, rng.max_row + 1):
                    for c in range(rng.min_col, rng.max_col + 1):
                        ws.cell(r, c).value = tl
        except Exception as ex:
            diffs.append(f"[{ws.title}] 結合セルの展開で例外: {ex}")

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        grid: List[List[str]] = []
        for r in range(1, max_row + 1):
            row: List[str] = []
            for c in range(1, max_col + 1):
                row.append(norm_text(ws.cell(r, c).value))
            grid.append(row)
        sheets[ws.title] = trim_grid(grid)

    if data_only:
        diffs.append("数式セルは保存済みの計算結果（cached value）しか取得できない場合があります（Excel側で未計算だと空になる可能性）。")
    return sheet_names, sheets, diffs


def extract_xlsx_images(path: Path, out_dir: Path) -> Tuple[Dict[str, List[str]], List[str]]:
    extracted: Dict[str, List[str]] = {}
    diffs: List[str] = []
    if load_workbook is None:
        return extracted, ["openpyxl 未導入のため .xlsx 画像抽出をスキップ。"]
    if Image is None:
        return extracted, ["pillow 未導入のため .xlsx 画像抽出をスキップ。"]

    wb = load_workbook(str(path), data_only=True)
    assets_dir = out_dir / (path.stem + "_assets") / "images"
    assets_dir.mkdir(parents=True, exist_ok=True)

    for ws in wb.worksheets:
        rels: List[str] = []
        imgs = getattr(ws, "_images", []) or []
        for idx, img in enumerate(imgs, start=1):
            try:
                data = img._data()
                im = Image.open(BytesIO(data))
                out_name = f"{slugify(ws.title)}_{idx:02d}.png"
                out_path = assets_dir / out_name
                im.save(out_path, format="PNG")
                rels.append(str(out_path.relative_to(out_dir)))
            except Exception as ex:
                diffs.append(f"[{ws.title}] 画像抽出に失敗: {ex}")
        if rels:
            extracted[ws.title] = rels

    return extracted, diffs


# -----------------------------
# DrawingML → Mermaid (xlsx)
# -----------------------------
_D_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def _zip_read(z: zipfile.ZipFile, name: str) -> Optional[bytes]:
    try:
        return z.read(name)
    except KeyError:
        return None


def map_sheetname_to_sheetxml(z: zipfile.ZipFile) -> Dict[str, str]:
    m: Dict[str, str] = {}
    wb_xml = _zip_read(z, "xl/workbook.xml")
    rels_xml = _zip_read(z, "xl/_rels/workbook.xml.rels")
    if not wb_xml or not rels_xml:
        return m

    wb_root = ET.fromstring(wb_xml)
    rels_root = ET.fromstring(rels_xml)

    rid_to_target: Dict[str, str] = {}
    for rel in rels_root.findall(".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
        rid = rel.get("Id")
        tgt = rel.get("Target")
        if rid and tgt:
            rid_to_target[rid] = tgt

    for sheet in wb_root.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
        name = sheet.get("name")
        rid = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if name and rid and rid in rid_to_target:
            target = rid_to_target[rid]
            if not target.startswith("xl/"):
                target = "xl/" + target.lstrip("/")
            m[name] = target
    return m


def map_sheetxml_to_drawing(z: zipfile.ZipFile, sheet_xml_path: str) -> Optional[str]:
    p = Path(sheet_xml_path)
    rels_path = str(p.parent / "_rels" / (p.name + ".rels"))
    rels_bytes = _zip_read(z, rels_path)
    if not rels_bytes:
        return None
    rels_root = ET.fromstring(rels_bytes)
    for rel in rels_root.findall(".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
        rtype = rel.get("Type")
        tgt = rel.get("Target")
        if rtype and "relationships/drawing" in rtype and tgt:
            base = p.parent
            resolved = (base / tgt).as_posix()
            resolved = str(Path(resolved)).replace("\\", "/")
            return resolved
    return None


@dataclass
class MermaidDiagram:
    mermaid: str
    metrics: Dict[str, Any]
    incomplete_connectors: List[Dict[str, Any]]


def drawingml_to_mermaid(drawing_xml: str) -> MermaidDiagram:
    root = ET.fromstring(drawing_xml)

    shapes: Dict[str, str] = {}
    x_positions: Dict[str, int] = {}
    colors: Dict[str, Optional[str]] = {}

    for anchor in root.findall(".//xdr:twoCellAnchor", _D_NS):
        sp = anchor.find(".//xdr:sp", _D_NS)
        if sp is None:
            continue
        nvSpPr = sp.find(".//xdr:nvSpPr", _D_NS)
        cNvPr = nvSpPr.find(".//xdr:cNvPr", _D_NS) if nvSpPr is not None else None
        if cNvPr is None:
            continue
        sid = cNvPr.get("id")
        if not sid:
            continue

        off = sp.find(".//a:xfrm/a:off", _D_NS)
        if off is not None and off.get("x") is not None:
            try:
                x_positions[sid] = int(off.get("x") or "0")
            except Exception:
                x_positions[sid] = 0

        text_parts = [t.text for t in sp.findall(".//a:t", _D_NS) if t.text]
        text = "".join(text_parts).strip()

        solid = sp.find(".//a:solidFill/a:srgbClr", _D_NS)
        color = solid.get("val") if solid is not None else None

        if text:
            shapes[sid] = text
            colors[sid] = color

    connections: List[Dict[str, Any]] = []
    incomplete: List[Dict[str, Any]] = []

    for anchor in root.findall(".//xdr:twoCellAnchor", _D_NS):
        cxnSp = anchor.find(".//xdr:cxnSp", _D_NS)
        if cxnSp is None:
            continue
        nvCxn = cxnSp.find(".//xdr:nvCxnSpPr", _D_NS)
        cNvCxn = nvCxn.find(".//xdr:cNvCxnSpPr", _D_NS) if nvCxn is not None else None
        st = cNvCxn.find(".//a:stCxn", _D_NS) if cNvCxn is not None else None
        ed = cNvCxn.find(".//a:endCxn", _D_NS) if cNvCxn is not None else None

        cNvPr = nvCxn.find(".//xdr:cNvPr", _D_NS) if nvCxn is not None else None
        cid = cNvPr.get("id") if cNvPr is not None else "unknown"

        if st is None or ed is None:
            incomplete.append({"connector_id": cid, "has_start": st is not None, "has_end": ed is not None})
            continue

        from_id = st.get("id")
        to_id = ed.get("id")
        if not from_id or not to_id:
            incomplete.append({"connector_id": cid, "has_start": bool(from_id), "has_end": bool(to_id)})
            continue

        ln = cxnSp.find(".//a:ln", _D_NS)
        line_color = None
        if ln is not None:
            clr = ln.find(".//a:solidFill/a:srgbClr", _D_NS)
            if clr is not None:
                line_color = clr.get("val")

        connections.append(
            {
                "from_id": from_id,
                "to_id": to_id,
                "from_text": shapes.get(from_id, f"Unknown({from_id})"),
                "to_text": shapes.get(to_id, f"Unknown({to_id})"),
                "line_color": line_color,
            }
        )

    metrics = {
        "total_shapes": len(shapes),
        "total_connectors": len(connections) + len(incomplete),
        "valid_connectors": len(connections),
        "incomplete_connectors": len(incomplete),
    }

    bucket_size = 2_000_000
    buckets: Dict[int, List[str]] = {}
    for sid in shapes.keys():
        x = x_positions.get(sid, 0)
        b = int(x // bucket_size)
        buckets.setdefault(b, []).append(sid)

    def node_id(sid: str) -> str:
        return f"N{sid}"

    color_to_class: Dict[str, str] = {}
    class_defs: List[str] = []
    class_assign: List[str] = []
    class_counter = 1
    for sid, clr in colors.items():
        if not clr:
            continue
        if clr not in color_to_class:
            cname = f"c{class_counter}"
            class_counter += 1
            color_to_class[clr] = cname
            class_defs.append(f"classDef {cname} fill:#{clr},stroke:#333,stroke-width:1px;")
        class_assign.append(f"class {node_id(sid)} {color_to_class[clr]};")

    lines: List[str] = ["flowchart LR"]
    for b in sorted(buckets.keys()):
        ids = buckets[b]
        lines.append(f"  subgraph col_{b}")
        for sid in ids:
            label = shapes.get(sid, "").replace('"', '\\"').replace("\n", " ")
            lines.append(f'    {node_id(sid)}["{label}"]')
        lines.append("  end")

    for c in connections:
        lines.append(f"  {node_id(c['from_id'])} --> {node_id(c['to_id'])}")

    if class_defs:
        lines.append("")
        lines.extend("  " + x for x in class_defs)
    if class_assign:
        lines.append("")
        lines.extend("  " + x for x in class_assign)

    return MermaidDiagram("\n".join(lines), metrics, incomplete)


def extract_xlsx_mermaid_per_sheet(path: Path) -> Tuple[Dict[str, MermaidDiagram], List[str]]:
    out: Dict[str, MermaidDiagram] = {}
    diffs: List[str] = []
    try:
        with zipfile.ZipFile(path, "r") as z:
            sheet_map = map_sheetname_to_sheetxml(z)
            if not sheet_map:
                return out, ["workbook.xml からシート対応が取れず、DrawingML→Mermaid をスキップしました。"]
            for sheet_name, sheet_xml in sheet_map.items():
                drawing_path = map_sheetxml_to_drawing(z, sheet_xml)
                if not drawing_path:
                    continue
                drawing_bytes = _zip_read(z, drawing_path)
                if not drawing_bytes:
                    diffs.append(f"[{sheet_name}] DrawingMLファイルを読み取れませんでした: {drawing_path}")
                    continue
                try:
                    diagram = drawingml_to_mermaid(drawing_bytes.decode("utf-8"))
                    if diagram.metrics.get("total_shapes", 0) > 0:
                        out[sheet_name] = diagram
                        if diagram.incomplete_connectors:
                            diffs.append(f"[{sheet_name}] DrawingMLコネクタに接続先不明が {len(diagram.incomplete_connectors)} 件あります。")
                except Exception as ex:
                    diffs.append(f"[{sheet_name}] DrawingML→Mermaid変換で例外: {ex}")
    except Exception as ex:
        diffs.append(f"xlsx を zip として扱う処理に失敗: {ex}")
    return out, diffs


# -----------------------------
# .xls reader
# -----------------------------
def _xlrd_cell_to_str(cell: Any, datemode: int) -> str:
    try:
        ctype = cell.ctype
        val = cell.value
    except Exception:
        return norm_text(cell)

    if ctype in (0, 6):
        return ""
    if ctype == 1:
        return str(val).strip()
    if ctype == 2:
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val)
    if ctype == 3:
        try:
            tup = xlrd.xldate_as_tuple(val, datemode)  # type: ignore[attr-defined]
            if tup[:3] == (0, 0, 0):
                return ""
            if tup[3:] == (0, 0, 0):
                return f"{tup[0]:04d}-{tup[1]:02d}-{tup[2]:02d}"
            return f"{tup[0]:04d}-{tup[1]:02d}-{tup[2]:02d}T{tup[3]:02d}:{tup[4]:02d}:{tup[5]:02d}"
        except Exception:
            return str(val)
    if ctype == 4:
        return "TRUE" if bool(val) else "FALSE"
    if ctype == 5:
        return f"#ERROR({val})"
    return str(val).strip()


def load_xls_grid(path: Path) -> Tuple[List[str], Dict[str, List[List[str]]], List[str]]:
    diffs: List[str] = []
    if xlrd is None:
        raise RuntimeError("xlrd is not installed. Please install: pip install xlrd==2.* (xls read)")
    wb = xlrd.open_workbook(str(path), formatting_info=True)
    sheet_names = wb.sheet_names()
    sheets: Dict[str, List[List[str]]] = {}
    datemode = wb.datemode

    for sh in wb.sheets():
        merged_map: Dict[Tuple[int, int], Any] = {}
        try:
            for (rlo, rhi, clo, chi) in getattr(sh, "merged_cells", []):
                v = sh.cell(rlo, clo).value
                for r in range(rlo, rhi):
                    for c in range(clo, chi):
                        merged_map[(r, c)] = v
        except Exception as ex:
            diffs.append(f"[{sh.name}] 結合セル情報の取得で例外: {ex}")

        grid: List[List[str]] = []
        for r in range(sh.nrows):
            row: List[str] = []
            for c in range(sh.ncols):
                if (r, c) in merged_map:
                    row.append(norm_text(merged_map[(r, c)]))
                else:
                    row.append(_xlrd_cell_to_str(sh.cell(r, c), datemode))
            grid.append(row)
        sheets[sh.name] = trim_grid(grid)

    diffs.append(".xls 形式は画像/図形の抽出が難しいため、本スクリプトでは画像抽出を行いません。")
    return sheet_names, sheets, diffs


# -----------------------------
# Logic template extraction
# -----------------------------
_LOGIC_KEYWORDS_SHEET = ("ロジック", "Logic", "BBR")
_LOGIC_HEADWORDS = [
    "機能概要", "ロジッククラス", "ロジックNO", "ステータス", "TX属性",
    "インタフェース定義", "パラメータ", "リターン",
    "データI/O定義", "データIO定義", "使用Dao名",
    "その他importクラス", "その他 import クラス",
    "コントラクト定義", "事前条件", "事後条件",
    "ロジック定義", "処理内容",
]


def _normalize_headword(s: str) -> str:
    s = s.strip()
    s = s.replace("データIO定義", "データI/O定義")
    s = s.replace("その他 import クラス", "その他importクラス")
    return s


def is_logic_template(sheet_name: str, grid: List[List[str]]) -> bool:
    cond1 = any(k in sheet_name for k in _LOGIC_KEYWORDS_SHEET)
    found = set()
    for r in grid:
        for c in r:
            t = norm_text(c)
            if not t:
                continue
            for hw in _LOGIC_HEADWORDS:
                if hw in t:
                    found.add(hw)
    cond2 = len(found) >= 3
    return cond1 and cond2


def find_cell(grid: List[List[str]], predicate) -> Optional[Tuple[int, int, str]]:
    for i, row in enumerate(grid):
        for j, cell in enumerate(row):
            t = norm_text(cell)
            if predicate(t):
                return (i, j, t)
    return None


def get_neighbor_value(grid: List[List[str]], i: int, j: int) -> str:
    if i < len(grid) and j + 1 < len(grid[i]):
        v = norm_text(grid[i][j + 1])
        if v:
            return v
    if i + 1 < len(grid) and j < len(grid[i + 1]):
        v = norm_text(grid[i + 1][j])
        if v:
            return v
    if i < len(grid):
        for jj in range(j + 1, len(grid[i])):
            v = norm_text(grid[i][jj])
            if v:
                return v
    return ""


def extract_logic_no_and_name(sheet_name: str, grid: List[List[str]]) -> Tuple[str, str]:
    logic_no = ""
    logic_name = ""

    hit = find_cell(grid, lambda t: t == "ロジックNO" or t.lower() == "logic no")
    if hit:
        i, j, _ = hit
        logic_no = get_neighbor_value(grid, i, j)

    if not logic_no:
        m = re.search(r"BBR\d+L\d+", sheet_name)
        if not m:
            all_text = " ".join(" ".join(r) for r in grid[:20])
            m = re.search(r"BBR\d+L\d+", all_text)
        logic_no = m.group(0) if m else "-"

    m2 = re.search(r"【\s*(BBR\d+L\d+)\s*】\s*(.+)$", sheet_name)
    if m2:
        logic_name = m2.group(2).strip()
    else:
        tmp = re.sub(r"BBR\d+L\d+", "", sheet_name)
        tmp = tmp.replace("【", "").replace("】", "").strip(" -_　")
        logic_name = tmp.strip() if tmp.strip() else "-"
    logic_name = re.sub(r"^LOGIC\s*", "", logic_name, flags=re.I).strip()
    logic_name = logic_name.replace("【", "").replace("】", "").strip() or "-"
    return logic_no or "-", logic_name


def extract_basic_info(grid: List[List[str]]) -> Dict[str, str]:
    keys = ["機能概要", "ロジッククラス", "ロジックNO", "ステータス", "TX属性"]
    out: Dict[str, str] = {}
    for k in keys:
        hit = find_cell(grid, lambda t, kk=k: _normalize_headword(t) == kk)
        if not hit:
            out[k] = "-"
            continue
        i, j, _ = hit
        v = get_neighbor_value(grid, i, j) or "-"
        if "[BIRD]" in v or "ロジック設計書" in v:
            v = "-"
        out[k] = v
    return out


def _header_map(header_cells: List[str]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for idx, h in enumerate(header_cells):
        t = norm_text(h)
        if not t:
            continue
        t = t.replace("　", " ").strip()
        t = t.replace("データ型", "データタイプ")
        if t == "型":
            t = "データタイプ"
        if t == "必須可否":
            t = "必須"
        if t == "説明":
            t = "内容"
        m[t] = idx
    return m


def extract_table_by_header(
    grid: List[List[str]],
    required_any: Sequence[str],
    until_headwords: Sequence[str],
    stop_blank_rows: int = 3,
) -> Optional[Tuple[int, int, List[List[str]]]]:
    norm_until = [_normalize_headword(x) for x in until_headwords]

    def row_has_headword(row: List[str]) -> bool:
        left = norm_text(row[0]) if row else ""
        if not left:
            return False
        return any(hw and hw in left for hw in norm_until)

    for i, row in enumerate(grid):
        header_cells = [norm_text(c) for c in row]
        header_set = set(h for h in header_cells if h)
        if not header_set:
            continue
        if not all(any(req in h for h in header_set) for req in required_any):
            continue

        rows: List[List[str]] = [row]
        blanks = 0
        j = i + 1
        while j < len(grid):
            r = grid[j]
            if row_has_headword(r) and j != i + 1:
                break
            if row_nonempty_count(r) == 0:
                blanks += 1
                if blanks >= stop_blank_rows:
                    break
            else:
                blanks = 0
                rows.append(r)
            j += 1
        return i, j, rows
    return None


def normalize_table_columns(rows: List[List[str]], desired_cols: List[str]) -> List[List[str]]:
    if not rows:
        return rows
    header = [norm_text(c) for c in rows[0]]
    hmap = _header_map(header)

    out_rows: List[List[str]] = [desired_cols]
    for r in rows[1:]:
        newr: List[str] = []
        for col in desired_cols:
            idx = None
            if col in hmap:
                idx = hmap[col]
            else:
                for k, v in hmap.items():
                    if col in k or k in col:
                        idx = v
                        break
            val = "-"
            if idx is not None and idx < len(r):
                v0 = norm_text(r[idx])
                val = v0 if v0 else "-"
            newr.append(val)
        if all(v == "-" for v in newr):
            continue
        out_rows.append(newr)
    return out_rows


_NUM_PAT = re.compile(r"^\s*([0-9]+|[０-９]+)\s*([\.．\)）]?)\s*(.*)\s*$")


def _zenkaku_to_hankaku_digits(s: str) -> str:
    trans = str.maketrans("０１２３４５６７８９", "0123456789")
    return s.translate(trans)


def wrap_constants_inline_code(s: str) -> str:
    def repl(m):
        txt = m.group(0)
        return f"`{txt}`"
    return re.sub(r"\{[^{}]{1,200}\}", repl, s)


def extract_processing_content(grid: List[List[str]], start_row: int) -> List[str]:
    out: List[str] = []
    i = start_row
    cur_step_open = False
    while i < len(grid):
        row = grid[i]
        left = norm_text(row[0]) if row else ""
        if left and any(hw in left for hw in ["インタフェース定義", "データI/O定義", "コントラクト定義"]):
            break

        cells = [norm_text(c) for c in row if norm_text(c)]
        if not cells:
            i += 1
            continue

        text = " ".join(cells)
        m = _NUM_PAT.match(text)
        if m:
            num = _zenkaku_to_hankaku_digits(m.group(1))
            rest = wrap_constants_inline_code(m.group(3).strip())
            out.append(f"{num}. **{rest or '（項目名未取得）'}**")
            cur_step_open = True
        else:
            c0 = norm_text(row[0])
            if _NUM_PAT.match(c0) and len(row) > 1:
                m0 = _NUM_PAT.match(c0)
                num = _zenkaku_to_hankaku_digits(m0.group(1))
                rest = wrap_constants_inline_code(" ".join(norm_text(x) for x in row[1:] if norm_text(x)).strip())
                out.append(f"{num}. **{rest or '（項目名未取得）'}**")
                cur_step_open = True
            else:
                desc = wrap_constants_inline_code(text)
                out.append(f"   - {desc}" if cur_step_open else desc)
        i += 1
    return out


def extract_logic_sections(sheet_name: str, grid: List[List[str]]) -> Tuple[str, List[str], List[str]]:
    diffs: List[str] = []
    unknowns: List[str] = []

    logic_no, logic_name = extract_logic_no_and_name(sheet_name, grid)
    md: List[str] = [f"## {logic_no}: {logic_name}", ""]

    basic = extract_basic_info(grid)
    md.append("### 基本情報")
    md.append("")
    md.append(md_table([["項目", "内容"]] + [[k, basic.get(k, "-")] for k in ["機能概要", "ロジッククラス", "ロジックNO", "ステータス", "TX属性"]]))
    md.append("")

    md.append("### インタフェース定義")
    md.append("")

    param_tbl = extract_table_by_header(grid, required_any=["項目名", "データ"], until_headwords=_LOGIC_HEADWORDS)
    if param_tbl:
        _, _, rows = param_tbl
        md.append("#### パラメータ")
        md.append("")
        md.append(md_table(normalize_table_columns(rows, ["項目名", "データタイプ", "内容", "必須", "初期値"])))
        md.append("")
    else:
        diffs.append(f"[{sheet_name}] パラメータ表の検出に失敗。")
        md.extend(["#### パラメータ", "", "_（抽出できませんでした）_", ""])

    md.append("#### リターン")
    md.append("")
    # Try to find another table after param
    ret_tbl = None
    if param_tbl:
        si, _, _ = param_tbl
        ret_tbl = extract_table_by_header(grid[si + 1 :], required_any=["項目名", "データ"], until_headwords=_LOGIC_HEADWORDS)
    if ret_tbl:
        _, _, rows = ret_tbl
        md.append(md_table(normalize_table_columns(rows, ["項目名", "データタイプ", "内容", "初期値"])))
        md.append("")
    else:
        md.append("_（抽出できませんでした）_")
        md.append("")
        unknowns.append(f"[{sheet_name}] リターン表が存在しない、または検出できませんでした。")

    md.append("### データI/O定義")
    md.append("")
    dio = extract_table_by_header(grid, required_any=["Dao", "使用"], until_headwords=_LOGIC_HEADWORDS)
    if dio:
        _, _, rows = dio
        md.append(md_table(normalize_table_columns(rows, ["使用Dao名", "Daoクラス名"])))
        md.append("")
    else:
        md.append("_（抽出できませんでした）_")
        md.append("")
        unknowns.append(f"[{sheet_name}] データI/O定義が存在しない、または検出できませんでした。")

    md.append("### その他importクラス")
    md.append("")
    imp = extract_table_by_header(grid, required_any=["クラス", "パッケージ"], until_headwords=_LOGIC_HEADWORDS)
    if imp:
        _, _, rows = imp
        md.append(md_table(normalize_table_columns(rows, ["クラス名", "パッケージ"])))
        md.append("")
    else:
        md.append("_（抽出できませんでした）_")
        md.append("")
        unknowns.append(f"[{sheet_name}] その他importクラスが存在しない、または検出できませんでした。")

    md.append("### コントラクト定義")
    md.append("")
    md.append("#### 事前条件")
    md.append("")
    pre = extract_table_by_header(grid, required_any=["条件", "メッセージ"], until_headwords=_LOGIC_HEADWORDS)
    if pre:
        _, _, rows = pre
        md.append(md_table(normalize_table_columns(rows, ["条件", "メッセージID（引数）"])))
        md.append("")
    else:
        md.append("_（抽出できませんでした）_")
        md.append("")
        unknowns.append(f"[{sheet_name}] 事前条件が存在しない、または検出できませんでした。")

    md.append("#### 事後条件")
    md.append("")
    post = None
    if pre:
        si, _, _ = pre
        post = extract_table_by_header(grid[si + 1 :], required_any=["条件", "メッセージ"], until_headwords=_LOGIC_HEADWORDS)
    if post:
        _, _, rows = post
        md.append(md_table(normalize_table_columns(rows, ["条件", "メッセージID（引数）"])))
        md.append("")
    else:
        md.append("_（抽出できませんでした）_")
        md.append("")
        unknowns.append(f"[{sheet_name}] 事後条件が存在しない、または検出できませんでした。")

    md.append("### ロジック定義")
    md.append("")
    md.append("#### 処理内容")
    md.append("")
    hit = find_cell(grid, lambda t: "処理内容" in t)
    if hit:
        i, _, _ = hit
        steps = extract_processing_content(grid, start_row=i + 1)
        if steps:
            md.extend(steps)
            md.append("")
        else:
            md.append("_（処理内容の抽出に失敗しました）_")
            md.append("")
            diffs.append(f"[{sheet_name}] 処理内容の抽出に失敗（番号付き行が検出できない）。")
    else:
        md.append("_（処理内容セクションを検出できませんでした）_")
        md.append("")
        diffs.append(f"[{sheet_name}] 「処理内容」見出しが見つかりません。")

    return "\n".join(md), diffs, unknowns


# -----------------------------
# Update history extraction
# -----------------------------
_UPDATE_SHEET_HINTS = ("更新履歴", "変更履歴", "改訂履歴")


def extract_update_history(grid: List[List[str]]) -> Optional[List[List[str]]]:
    for i, row in enumerate(grid):
        joined = " ".join(norm_text(c) for c in row if norm_text(c))
        if "作成" in joined and "更新" in joined and ("更新内容" in joined or "内容" in joined):
            start = i
            blanks = 0
            rows: List[List[str]] = []
            j = start
            while j < len(grid):
                r = grid[j]
                if row_nonempty_count(r) == 0:
                    blanks += 1
                    if blanks >= 2:
                        break
                else:
                    blanks = 0
                    rows.append(r)
                j += 1
            desired = ["作成・更新日", "更新内容", "作成・更新者", "レビュー/承認者", "レビュー/承認日"]
            return normalize_table_columns(rows, desired)
    return None


def extract_document_info(title: str, file_name: str, sheet_names: Sequence[str], grids: Dict[str, List[List[str]]]) -> Dict[str, str]:
    info = {
        "管理番号": detect_management_no(file_name) or "-",
        "分類": guess_doc_classification(title, sheet_names),
        "作成者": "-",
        "作成日": "-",
        "更新者": "-",
        "更新日": "-",
    }
    # try from update sheet labels
    for sname in sheet_names:
        if any(h in sname for h in _UPDATE_SHEET_HINTS):
            g = grids.get(sname, [])
            for lab in ["作成者", "作成日", "更新者", "更新日", "分類", "管理番号"]:
                hit = find_cell(g, lambda t, L=lab: t == L)
                if hit:
                    i, j, _ = hit
                    v = get_neighbor_value(g, i, j)
                    if v:
                        if lab in info:
                            info[lab] = v
                        elif lab == "管理番号":
                            info["管理番号"] = v
                        elif lab == "分類":
                            info["分類"] = v
            break
    return info


def build_overview(title: str, sheet_names: Sequence[str]) -> str:
    cls = guess_doc_classification(title, sheet_names)
    lines = [f"- 分類推定: {cls}", f"- シート数: {len(sheet_names)}"]
    for s in sheet_names[:30]:
        hint = ""
        if any(h in s for h in _UPDATE_SHEET_HINTS):
            hint = "（更新履歴）"
        elif any(k in s for k in _LOGIC_KEYWORDS_SHEET):
            hint = "（ロジック）"
        lines.append(f"  - {s}{hint}")
    if len(sheet_names) > 30:
        lines.append(f"  - ...（省略 {len(sheet_names)-30} 件）")
    return "\n".join(lines)


def convert_sheet_generic(sheet_name: str, grid: List[List[str]]) -> str:
    parts: List[str] = [f"## {sheet_name}", ""]
    if not grid:
        parts.append("_（空シート）_")
        parts.append("")
        return "\n".join(parts)

    blocks = split_blocks_by_empty_rows(grid, empty_run=2)
    for bi, block in enumerate(blocks, start=1):
        if not block:
            continue
        if is_table_block(block):
            h = find_actual_header_row(block, sheet_name)
            block2 = block[h:] if h < len(block) else block
            block2 = [clean_header_row([norm_text(x) for x in block2[0]])] + block2[1:]
            parts.extend([f"### Table {bi}", "", md_table(block2), ""])
        else:
            parts.extend([f"### Notes {bi}", "", md_paragraph_block(block), ""])
    return "\n".join(parts)


def convert_workbook_to_markdown(
    in_path: Path,
    out_dir: Optional[Path] = None,
    data_only: bool = True,
    include_images: bool = True,
    include_mermaid: bool = True,
) -> Tuple[str, List[str], List[str], Path]:
    difficulties: List[str] = []
    unknowns: List[str] = []

    in_path = in_path.resolve()
    out_dir = in_path.parent if out_dir is None else out_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    title = in_path.stem
    ext = in_path.suffix.lower()

    images_map: Dict[str, List[str]] = {}
    mermaid_map: Dict[str, MermaidDiagram] = {}

    if ext == ".xlsx":
        sheet_names, grids, dif0 = load_xlsx_grid(in_path, data_only=data_only)
        difficulties.extend(dif0)
        if include_images:
            images_map, dif_img = extract_xlsx_images(in_path, out_dir)
            difficulties.extend(dif_img)
        if include_mermaid:
            mermaid_map, dif_m = extract_xlsx_mermaid_per_sheet(in_path)
            difficulties.extend(dif_m)
    elif ext == ".xls":
        sheet_names, grids, dif0 = load_xls_grid(in_path)
        difficulties.extend(dif0)
    else:
        raise ValueError(f"Unsupported extension: {ext} (need .xls/.xlsx)")

    doc_info = extract_document_info(title, in_path.name, sheet_names, grids)

    update_history_rows: Optional[List[List[str]]] = None
    for sname in sheet_names:
        if any(h in sname for h in _UPDATE_SHEET_HINTS):
            update_history_rows = extract_update_history(grids.get(sname, []))
            break

    md: List[str] = []
    md.append(f"# {title}")
    md.append("")
    md.append("## 概要")
    md.append("")
    md.append(build_overview(title, sheet_names))
    md.append("")
    md.append("## 文書情報")
    md.append("")
    md.append(md_table([["項目", "内容"]] + [[k, doc_info.get(k, "-")] for k in ["管理番号", "分類", "作成者", "作成日", "更新者", "更新日"]]))
    md.append("")
    md.append("## 更新履歴")
    md.append("")
    if update_history_rows:
        md.append(md_table(update_history_rows))
    else:
        md.append(md_table([["作成・更新日", "更新内容", "作成・更新者", "レビュー/承認者", "レビュー/承認日"], ["-", "-", "-", "-", "-"]]))
        unknowns.append("更新履歴シート（更新履歴/変更履歴）が見つからない、または表の抽出に失敗しました。")
    md.append("")
    md.append("---")
    md.append("")

    for sname in sheet_names:
        grid = grids.get(sname, [])
        if is_logic_template(sname, grid):
            try:
                sect_md, dif_s, unk_s = extract_logic_sections(sname, grid)
                md.append(sect_md)
                difficulties.extend(dif_s)
                unknowns.extend(unk_s)
            except Exception as ex:
                difficulties.append(f"[{sname}] ロジック抽出で例外: {ex} → フォールバック出力。")
                md.append(convert_sheet_generic(sname, grid))
        else:
            md.append(convert_sheet_generic(sname, grid))

        if include_images and ext == ".xlsx":
            imgs = images_map.get(sname, [])
            if imgs:
                md.append("### 画像")
                md.append("")
                for rel in imgs:
                    md.append(f"![{sname}]({rel})")
                md.append("")

        if include_mermaid and ext == ".xlsx":
            diag = mermaid_map.get(sname)
            if diag:
                md.append("### 図形 (Mermaid 推定)")
                md.append("")
                md.append(f"- 変換メトリクス: shapes={diag.metrics.get('total_shapes')}, connectors={diag.metrics.get('total_connectors')}, valid={diag.metrics.get('valid_connectors')}, incomplete={diag.metrics.get('incomplete_connectors')}")
                md.append("")
                md.append("```mermaid")
                md.append(diag.mermaid)
                md.append("```")
                md.append("")
                if diag.incomplete_connectors:
                    difficulties.append(f"[{sname}] Mermaid 変換で接続先不明コネクタ {len(diag.incomplete_connectors)} 件。")

        md.append("---")
        md.append("")

    md.append("## 読み取りが難しかった項目")
    md.append("")
    if difficulties:
        for d in sorted(set(difficulties)):
            md.append(f"- {d}")
    else:
        md.append("- （特になし）")
    md.append("")
    md.append("## 不明点・不明瞭な点")
    md.append("")
    if unknowns:
        for u in sorted(set(unknowns)):
            md.append(f"- {u}")
    else:
        md.append("- （特になし）")
    md.append("")

    md_text = "\n".join(md)
    out_md = out_dir / f"{in_path.stem}.md"
    out_md.write_text(md_text, encoding="utf-8")
    return md_text, difficulties, unknowns, out_md


def main(argv: Optional[Sequence[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Convert Excel (.xls/.xlsx) into Markdown with per-sheet sections.")
    p.add_argument("excel_path", help="Path to .xls or .xlsx file")
    p.add_argument("-o", "--out-dir", default=None, help="Output directory (default: same as input)")
    p.add_argument("--no-images", action="store_true", help="Disable image extraction (.xlsx only)")
    p.add_argument("--no-mermaid", action="store_true", help="Disable DrawingML→Mermaid (.xlsx only)")
    p.add_argument("--raw-formulas", action="store_true", help="For .xlsx: do NOT use data_only (might show formulas)")
    p.add_argument("-v", "--verbose", action="store_true", help="Verbose logs")
    args = p.parse_args(argv)

    in_path = Path(args.excel_path)
    if not in_path.exists():
        eprint(f"ERROR: file not found: {in_path}")
        return 2

    out_dir = Path(args.out_dir) if args.out_dir else None
    data_only = not args.raw_formulas
    include_images = not args.no_images
    include_mermaid = not args.no_mermaid

    try:
        _, difficulties, unknowns, out_md = convert_workbook_to_markdown(
            in_path,
            out_dir=out_dir,
            data_only=data_only,
            include_images=include_images,
            include_mermaid=include_mermaid,
        )
        if args.verbose:
            eprint(f"Saved: {out_md}")
            eprint(f"Difficulties: {len(difficulties)}  Unknowns: {len(unknowns)}")
        else:
            print(str(out_md))
        return 0
    except Exception as ex:
        eprint(f"ERROR: {ex}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
